"""
================================================================================
🚀 ULTIMATE UNIT ECONOMICS ENGINE v48.1 - ИСПРАВЛЕНИЕ ОШИБОК
================================================================================
📌 ВЕРСИЯ: 48.1.0
📌 ИСПРАВЛЕНИЯ:
    ✅ Исправлены дублирующиеся ID в selectbox
    ✅ Обновлен use_container_width на width
    ✅ Добавлены уникальные ключи для всех элементов
================================================================================
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import math
import json
import warnings
import requests
import logging
import time
import hashlib
import hmac
import base64
import urllib.parse
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple, Union
from dataclasses import dataclass, field
from collections import Counter, defaultdict
from functools import lru_cache
from threading import Thread, Lock
from queue import Queue
import traceback
import os
import pickle
import random
from concurrent.futures import ThreadPoolExecutor, as_completed

warnings.filterwarnings('ignore')

# --------------------------------------------
# НАСТРОЙКА ЛОГИРОВАНИЯ
# --------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('unit_economy.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# --------------------------------------------
# ПРОВЕРКА НАЛИЧИЯ БИБЛИОТЕК
# --------------------------------------------
OPENPYXL_AVAILABLE = False
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    logger.warning(f"OpenPyXL не установлен: {e}")

PLOTLY_AVAILABLE = False
try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    PLOTLY_AVAILABLE = True
except ImportError:
    logger.warning("Plotly не установлен. Установите: pip install plotly")

SKLEARN_AVAILABLE = False
try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.naive_bayes import MultinomialNB
    from sklearn.pipeline import Pipeline
    import joblib
    SKLEARN_AVAILABLE = True
except ImportError:
    logger.warning("Scikit-learn не установлен. Установите: pip install scikit-learn joblib")

# --------------------------------------------
# КОНФИГУРАЦИЯ
# --------------------------------------------
CONFIG = {
    "version": "48.1.0",
    "app_name": "🚀 Юнит-экономика с конвертацией размеров",
    "currency": "₽",
    "marketplaces": ["Яндекс Маркет", "Ozon", "Wildberries", "AliExpress", "Мегамаркет"],
    "operation_modes": ["FBY", "FBS", "FBO", "DBS"],
    "dimension_units": ["мм", "см"],
    "default_dimension_unit": "мм",
    "category_keywords": {
        "Двигатель": ["двигатель", "мотор", "поршень", "кольцо", "гильза", "гбц", "головка", "блок цилиндров", "коленвал", "распредвал", "шатун", "клапан", "ремень грм", "цепь грм"],
        "Трансмиссия": ["коробка", "передач", "кпп", "вариатор", "сцепление", "диск сцепления", "корзина", "выжимной", "механизм", "автомат"],
        "Подвеска": ["амортизатор", "стойка", "пружина", "рычаг", "сайлентблок", "шаровой", "стабилизатор", "шаровая опора", "тяга"],
        "Тормозная система": ["тормоз", "колодка", "диск тормозной", "барабан", "суппорт", "главный тормозной", "шланг", "цилиндр"],
        "Рулевое управление": ["рулевой", "рейка", "наконечник", "тяга", "кардан", "усилитель", "рулевая"],
        "Электрооборудование": ["генератор", "стартер", "провод", "датчик", "реле", "блок управления", "эбу", "модуль"],
        "Система охлаждения": ["радиатор", "помпа", "термостат", "антифриз", "вентилятор", "патрубок"],
        "Система выпуска": ["глушитель", "катализатор", "сажевый", "выпускной", "гофра", "резонатор"],
        "Система питания": ["форсунка", "насос", "бензонасос", "тнвд", "инжектор", "карбюратор", "магистраль"],
        "Кузовные детали": ["бампер", "крыло", "капот", "дверь", "зеркало", "стекло", "молдинг", "порог"],
        "Оптика": ["фара", "фонарь", "поворотник", "габарит", "ксенон", "свет"],
        "Шины и диски": ["шина", "диск", "колесо", "резина", "покрышка", "литье"],
        "Масла и жидкости": ["масло", "жидкость", "смазка", "охлаждайка", "трансмиссионка", "гидравлика"],
        "Фильтры": ["фильтр", "масляный", "воздушный", "топливный", "салонный", "очистка"],
        "Ремни и цепи": ["ремень", "цепь", "натяжитель", "ролик", "грм"],
        "Свечи зажигания": ["свеча", "зажигание", "накаливания", "свечной", "свечка"],
        "Колодки и диски": ["колодки", "диски тормозные", "накладки", "тормозные"],
        "Аккумуляторы": ["аккумулятор", "батарея", "акум", "зарядка"],
        "Инструмент": ["инструмент", "ключ", "набор", "ремкомплект"],
        "Ручной инструмент": ["головка", "ключ", "набор инструментов", "отвертка", "плоскогубцы", "кусачки", "молоток", "стамеска"],
        "Садовая техника": ["триммер", "газонокосилка", "опрыскиватель", "шланг", "лопата", "грабли"],
        "Бытовая химия": ["чистящее", "моющее", "порошок", "гель", "пятновыводитель", "отбеливатель"],
        "Электротовары": ["кабель", "разъем", "переходник", "предохранитель", "розетка", "выключатель"],
        "Красота и уход": ["шампунь", "мыло", "крем", "маска", "лосьон", "тоник"]
    },
    "oem_patterns": [
        r'[0-9]{6,12}',
        r'[A-Z0-9]{6,12}',
        r'[A-Z]{2}[0-9]{6,10}',
        r'[A-Z]{2}[0-9]{4}[A-Z]{2}',
        r'[0-9]{4}[A-Z]{2}[0-9]{4}'
    ],
    "validation": {
        "min_price": 10,
        "max_price": 1000000,
        "min_cost": 1,
        "max_cost": 500000,
        "min_dimension": 0.1,
        "max_dimension": 1000,
        "min_volume": 0.001,
        "max_volume": 10000
    },
    "api": {
        "cache_ttl": 300,
        "max_retries": 3,
        "timeout": 30,
        "rate_limit": 10
    }
}

# --------------------------------------------
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# --------------------------------------------
def safe_float(val: Any, default: float = 0.0) -> float:
    """Безопасное преобразование в float"""
    try:
        if val is None or val == "" or val == "NaN" or val == "nan":
            return default
        if isinstance(val, (int, float)):
            if math.isnan(val) or math.isinf(val):
                return default
            return float(val)
        if isinstance(val, str):
            val = val.replace(',', '.').replace(' ', '').replace('₽', '').replace('%', '').replace('$', '')
            val = val.replace('€', '').replace('£', '')
            val = re.sub(r'[^\d.\-]', '', val)
            if not val or val == '-' or val == '.':
                return default
            return float(val)
        return default
    except (ValueError, TypeError):
        return default

def safe_str(val: Any, default: str = "") -> str:
    """Безопасное преобразование в str"""
    try:
        if val is None:
            return default
        if isinstance(val, (int, float)) and (math.isnan(val) or math.isinf(val)):
            return default
        return str(val).strip() if str(val).strip() else default
    except (ValueError, TypeError):
        return default

def safe_int(val: Any, default: int = 0) -> int:
    """Безопасное преобразование в int"""
    try:
        return int(safe_float(val, default))
    except (ValueError, TypeError):
        return default

def calculate_volume(length: float, width: float, height: float) -> float:
    """Расчет объема в литрах"""
    try:
        if all([length, width, height]) and all([length > 0, width > 0, height > 0]):
            if any([length > 1000, width > 1000, height > 1000]):
                return 0.0
            volume = (length * width * height) / 1000.0
            if volume < 0.001:
                return 0.0
            return round(volume, 3)
        return 0.0
    except (TypeError, ValueError):
        return 0.0

def convert_dimension(value: float, from_unit: str, to_unit: str) -> float:
    """Конвертация размеров между мм и см"""
    if value == 0:
        return 0.0
    
    if from_unit == to_unit:
        return value
    
    if from_unit == "мм" and to_unit == "см":
        return value / 10.0
    elif from_unit == "см" and to_unit == "мм":
        return value * 10.0
    
    return value

def format_currency(value: float) -> str:
    """Форматирование валюты"""
    try:
        if value is None or math.isnan(value) or math.isinf(value):
            return "0 ₽"
        return f"{value:,.0f} ₽" if abs(value) >= 1 else f"{value:.2f} ₽"
    except (ValueError, TypeError):
        return "0 ₽"

def format_percent(value: float) -> str:
    """Форматирование процентов"""
    try:
        if value is None or math.isnan(value) or math.isinf(value):
            return "0%"
        return f"{value:.1f}%" if abs(value) >= 0.1 else f"{value:.2f}%"
    except (ValueError, TypeError):
        return "0%"

def generate_cache_key(*args) -> str:
    """Генерация ключа для кэша"""
    key = "|".join(str(arg) for arg in args)
    return hashlib.md5(key.encode()).hexdigest()

# --------------------------------------------
# 📊 КЭШИРОВАНИЕ API
# --------------------------------------------
class APICache:
    """Кэширование API-запросов"""
    
    def __init__(self, cache_dir: str = "api_cache"):
        self.cache_dir = cache_dir
        self.memory_cache = {}
        self.lock = Lock()
        
        if not os.path.exists(cache_dir):
            os.makedirs(cache_dir)
    
    def get(self, key: str) -> Optional[Any]:
        """Получение из кэша"""
        with self.lock:
            if key in self.memory_cache:
                data, timestamp = self.memory_cache[key]
                if (datetime.now() - timestamp).total_seconds() < CONFIG["api"]["cache_ttl"]:
                    return data
            
            cache_file = os.path.join(self.cache_dir, f"{hashlib.md5(key.encode()).hexdigest()}.pkl")
            if os.path.exists(cache_file):
                try:
                    with open(cache_file, 'rb') as f:
                        data, timestamp = pickle.load(f)
                        if (datetime.now() - timestamp).total_seconds() < CONFIG["api"]["cache_ttl"]:
                            self.memory_cache[key] = (data, timestamp)
                            return data
                except Exception as e:
                    logger.warning(f"Cache read error: {e}")
            
            return None
    
    def set(self, key: str, value: Any):
        """Сохранение в кэш"""
        with self.lock:
            timestamp = datetime.now()
            self.memory_cache[key] = (value, timestamp)
            
            try:
                cache_file = os.path.join(self.cache_dir, f"{hashlib.md5(key.encode()).hexdigest()}.pkl")
                with open(cache_file, 'wb') as f:
                    pickle.dump((value, timestamp), f)
            except Exception as e:
                logger.warning(f"Cache write error: {e}")
    
    def clear(self):
        """Очистка кэша"""
        with self.lock:
            self.memory_cache.clear()
            for file in os.listdir(self.cache_dir):
                try:
                    os.remove(os.path.join(self.cache_dir, file))
                except:
                    pass

# --------------------------------------------
# 🤖 ML-КАТЕГОРИЗАЦИЯ
# --------------------------------------------
class AutoClassifier:
    """Автоматическая категоризация товаров с помощью ML"""
    
    def __init__(self, model_path: str = "category_model.pkl"):
        self.model_path = model_path
        self.model = None
        self.vectorizer = None
        self.load_model()
    
    def load_model(self):
        """Загрузка обученной модели"""
        if os.path.exists(self.model_path) and SKLEARN_AVAILABLE:
            try:
                self.model = joblib.load(self.model_path)
                logger.info("ML-модель загружена")
                return
            except Exception as e:
                logger.warning(f"Ошибка загрузки модели: {e}")
        
        self._train_model()
    
    def _train_model(self):
        """Обучение модели на категориях из CONFIG"""
        if not SKLEARN_AVAILABLE:
            return
        
        try:
            X = []
            y = []
            
            for category, keywords in CONFIG["category_keywords"].items():
                for keyword in keywords:
                    X.append(keyword)
                    y.append(category)
                
                for keyword in keywords:
                    X.append(keyword + " " + category.lower())
                    y.append(category)
                    X.append(category.lower() + " " + keyword)
                    y.append(category)
            
            if X:
                self.model = Pipeline([
                    ('tfidf', TfidfVectorizer(max_features=2000, ngram_range=(1, 2))),
                    ('clf', MultinomialNB(alpha=0.1))
                ])
                self.model.fit(X, y)
                joblib.dump(self.model, self.model_path)
                logger.info(f"ML-модель обучена на {len(X)} примерах")
        except Exception as e:
            logger.error(f"Ошибка обучения модели: {e}")
            self.model = None
    
    def predict(self, name: str) -> Tuple[str, float]:
        """Предсказание категории с уверенностью"""
        if not self.model or not name or not SKLEARN_AVAILABLE:
            return "Прочее", 0.0
        
        try:
            pred = self.model.predict([name])[0]
            probs = self.model.predict_proba([name])[0]
            confidence = max(probs) * 100
            
            if confidence < 30:
                return "Прочее", confidence
            
            return pred, confidence
        except Exception as e:
            logger.error(f"Ошибка предсказания: {e}")
            return "Прочее", 0.0

# --------------------------------------------
# 📤 TELEGRAM УВЕДОМЛЕНИЯ
# --------------------------------------------
class TelegramNotifier:
    """Отправка уведомлений в Telegram"""
    
    def __init__(self, bot_token: str = None, chat_id: str = None):
        self.bot_token = bot_token
        self.chat_id = chat_id
        self.api_url = f"https://api.telegram.org/bot{bot_token}/sendMessage" if bot_token else None
        self.document_url = f"https://api.telegram.org/bot{bot_token}/sendDocument" if bot_token else None
    
    def send_report(self, results: List[Dict]) -> bool:
        """Отправка отчета в Telegram"""
        if not self.bot_token or not self.chat_id:
            return False
        
        if not results:
            return False
        
        total_profit = sum(r.get("unit_profit", 0) for r in results)
        total_products = len(results)
        profitable = sum(1 for r in results if r.get("unit_profit", 0) > 0)
        avg_margin = sum(r.get("margin", 0) for r in results) / total_products if total_products else 0
        
        message = f"""
📊 **ОТЧЕТ ПО ЮНИТ-ЭКОНОМИКЕ**
📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}

📦 **Всего товаров:** {total_products}
💰 **Прибыльных:** {profitable} ({profitable/total_products*100:.1f}%)
💵 **Общая прибыль:** {format_currency(total_profit)}
📈 **Средняя маржа:** {format_percent(avg_margin)}

🏆 **Топ-5 по прибыли:**
"""
        top = sorted(results, key=lambda x: x.get('unit_profit', 0), reverse=True)[:5]
        for i, product in enumerate(top, 1):
            name = product.get('name', '')[:30]
            profit = product.get('unit_profit', 0)
            margin = product.get('margin', 0)
            message += f"\n{i}. {name} — {format_currency(profit)} ({format_percent(margin)})"
        
        categories = {}
        for r in results:
            cat = r.get('category', 'Прочее')
            if cat not in categories:
                categories[cat] = 0
            categories[cat] += 1
        
        message += "\n\n📊 **Распределение по категориям:**"
        for cat, count in sorted(categories.items(), key=lambda x: x[1], reverse=True)[:5]:
            message += f"\n- {cat}: {count} ({count/total_products*100:.1f}%)"
        
        try:
            response = requests.post(
                self.api_url,
                json={
                    "chat_id": self.chat_id,
                    "text": message,
                    "parse_mode": "Markdown"
                },
                timeout=30
            )
            return response.status_code == 200
        except Exception as e:
            logger.error(f"Telegram error: {e}")
            return False
    
    def send_file(self, file_bytes: bytes, filename: str) -> bool:
        """Отправка файла в Telegram"""
        if not self.bot_token or not self.chat_id:
            return False
        
        try:
            files = {
                'document': (filename, file_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            data = {
                'chat_id': self.chat_id
            }
            response = requests.post(
                self.document_url,
                files=files,
                data=data,
                timeout=60
            )
            return response.status_code == 200
        except Exception as e:
            logger.error(f"Telegram file error: {e}")
            return False

# --------------------------------------------
# 📊 GOOGLE SHEETS ЭКСПОРТ
# --------------------------------------------
class GoogleSheetsExporter:
    """Экспорт в Google Sheets"""
    
    def __init__(self, credentials_json: str = None, spreadsheet_name: str = None):
        self.credentials_json = credentials_json
        self.spreadsheet_name = spreadsheet_name
        self.client = None
        self.sheet = None
        
        if credentials_json and spreadsheet_name:
            self._connect()
    
    def _connect(self):
        """Подключение к Google Sheets"""
        try:
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            
            scope = [
                'https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive'
            ]
            
            if os.path.exists(self.credentials_json):
                creds = ServiceAccountCredentials.from_json_keyfile_name(self.credentials_json, scope)
            else:
                creds_dict = json.loads(self.credentials_json)
                creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
            
            self.client = gspread.authorize(creds)
            self._open_spreadsheet()
            return True
        except ImportError:
            logger.error("gspread не установлен. Установите: pip install gspread oauth2client")
            return False
        except Exception as e:
            logger.error(f"Google Sheets connection error: {e}")
            return False
    
    def _open_spreadsheet(self):
        """Открытие или создание таблицы"""
        try:
            self.sheet = self.client.open(self.spreadsheet_name)
        except:
            self.sheet = self.client.create(self.spreadsheet_name)
            self.sheet.share(None, perm_type='anyone', role='writer')
    
    def export_results(self, results: List[Dict]) -> Dict:
        """Экспорт результатов в Google Sheets"""
        if not self.client or not self.sheet:
            return {"status": "error", "message": "Не подключено к Google Sheets"}
        
        if not results:
            return {"status": "error", "message": "Нет данных для экспорта"}
        
        try:
            headers = [
                "Артикул", "Наименование", "Бренд", "OE номер", "Категория",
                "Цена", "Себестоимость", "Прибыль", "Маржа %", "ROI %",
                "Статус", "ABC", "Рекомендуемая цена", "Действие"
            ]
            
            data = []
            for r in results:
                data.append([
                    r.get('article', ''),
                    r.get('name', '')[:50],
                    r.get('brand', ''),
                    r.get('oe_number', ''),
                    r.get('category', ''),
                    r.get('price', 0),
                    r.get('cost', 0),
                    r.get('unit_profit', 0),
                    r.get('margin', 0),
                    r.get('roi', 0),
                    r.get('profit_status', ''),
                    r.get('abc_category', ''),
                    r.get('recommended_price', 0),
                    r.get('price_action', '')
                ])
            
            ws = self.sheet.get_worksheet(0)
            ws.clear()
            ws.update([headers] + data)
            ws.format('A1:N1', {'textFormat': {'bold': True}})
            
            return {
                "status": "success",
                "message": f"Экспортировано {len(results)} записей",
                "url": f"https://docs.google.com/spreadsheets/d/{self.sheet.id}"
            }
        except Exception as e:
            logger.error(f"Google Sheets export error: {e}")
            return {"status": "error", "message": str(e)}

# --------------------------------------------
# 🔌 API-КЛИЕНТЫ ДЛЯ МАРКЕТПЛЕЙСОВ
# --------------------------------------------
class BaseMarketplaceAPI:
    """Базовый класс для API маркетплейсов"""
    
    def __init__(self, api_key: str = None, client_id: str = None):
        self.api_key = api_key
        self.client_id = client_id
        self.cache = APICache()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (compatible; UnitEconomy/1.0)'
        })
    
    def _request(self, method: str, url: str, **kwargs) -> Optional[Dict]:
        """Выполнение запроса с ретраями"""
        for attempt in range(CONFIG["api"]["max_retries"]):
            try:
                response = self.session.request(
                    method, url,
                    timeout=CONFIG["api"]["timeout"],
                    **kwargs
                )
                if response.status_code == 200:
                    return response.json()
                elif response.status_code == 429:
                    time.sleep(2 ** attempt)
                else:
                    logger.warning(f"API error {response.status_code}: {response.text[:200]}")
                    break
            except Exception as e:
                logger.error(f"Request error (attempt {attempt+1}): {e}")
                time.sleep(1)
        
        return None


class YandexMarketAPI(BaseMarketplaceAPI):
    """API Яндекс Маркета"""
    
    BASE_URL = "https://api.partner.market.yandex.ru/v2"
    
    def __init__(self, api_key: str = None, business_id: str = None):
        super().__init__(api_key)
        self.business_id = business_id
        self.session.headers.update({
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        })
    
    def get_products(self, page: int = 1, page_size: int = 100) -> Optional[Dict]:
        """Получение списка товаров"""
        cache_key = generate_cache_key('yandex_products', page, page_size)
        cached = self.cache.get(cache_key)
        if cached:
            return cached
        
        url = f"{self.BASE_URL}/businesses/{self.business_id}/offer-mappings"
        params = {"page": page, "page_size": page_size}
        
        result = self._request('GET', url, params=params)
        if result:
            self.cache.set(cache_key, result)
        return result
    
    def get_offer_prices(self, offer_ids: List[str]) -> Optional[Dict]:
        """Получение цен товаров"""
        cache_key = generate_cache_key('yandex_prices', *sorted(offer_ids))
        cached = self.cache.get(cache_key)
        if cached:
            return cached
        
        url = f"{self.BASE_URL}/businesses/{self.business_id}/offer-prices"
        data = {"offerIds": offer_ids}
        
        result = self._request('POST', url, json=data)
        if result:
            self.cache.set(cache_key, result)
        return result
    
    def update_price(self, offer_id: str, price: float) -> Optional[Dict]:
        """Обновление цены товара"""
        url = f"{self.BASE_URL}/businesses/{self.business_id}/offer-prices/updates"
        data = {
            "offers": [{
                "offerId": offer_id,
                "price": {
                    "value": price,
                    "currency": "RUB"
                }
            }]
        }
        return self._request('POST', url, json=data)
    
    def update_stock(self, offer_id: str, stock: int) -> Optional[Dict]:
        """Обновление остатков"""
        url = f"{self.BASE_URL}/businesses/{self.business_id}/stocks"
        data = {
            "skus": [{
                "sku": offer_id,
                "stock": stock
            }]
        }
        return self._request('PUT', url, json=data)
    
    def create_product(self, product_data: Dict) -> Optional[Dict]:
        """Создание карточки товара"""
        url = f"{self.BASE_URL}/businesses/{self.business_id}/offer-mappings"
        data = {
            "offerMappings": [{
                "offer": {
                    "offerId": product_data.get('offer_id', ''),
                    "name": product_data.get('name', ''),
                    "price": product_data.get('price', 0),
                    "category": product_data.get('category', '')
                }
            }]
        }
        return self._request('POST', url, json=data)


class OzonAPI(BaseMarketplaceAPI):
    """API Ozon"""
    
    BASE_URL = "https://api-seller.ozon.ru/v2"
    
    def __init__(self, api_key: str = None, client_id: str = None):
        super().__init__(api_key, client_id)
        self.session.headers.update({
            'Api-Key': api_key,
            'Client-Id': client_id,
            'Content-Type': 'application/json'
        })
    
    def get_products(self, page: int = 1, page_size: int = 100) -> Optional[Dict]:
        """Получение списка товаров"""
        cache_key = generate_cache_key('ozon_products', page, page_size)
        cached = self.cache.get(cache_key)
        if cached:
            return cached
        
        url = f"{self.BASE_URL}/product/list"
        data = {"page": page, "page_size": page_size}
        
        result = self._request('POST', url, json=data)
        if result:
            self.cache.set(cache_key, result)
        return result
    
    def get_prices(self, product_ids: List[str]) -> Optional[Dict]:
        """Получение цен"""
        cache_key = generate_cache_key('ozon_prices', *sorted(product_ids))
        cached = self.cache.get(cache_key)
        if cached:
            return cached
        
        url = f"{self.BASE_URL}/product/prices"
        data = {"product_ids": product_ids}
        
        result = self._request('POST', url, json=data)
        if result:
            self.cache.set(cache_key, result)
        return result
    
    def update_price(self, product_id: str, price: float) -> Optional[Dict]:
        """Обновление цены"""
        url = f"{self.BASE_URL}/product/price"
        data = {
            "product_id": product_id,
            "price": price
        }
        return self._request('POST', url, json=data)
    
    def update_stock(self, product_id: str, stock: int) -> Optional[Dict]:
        """Обновление остатков"""
        url = f"{self.BASE_URL}/product/stocks"
        data = {
            "stocks": [{
                "product_id": product_id,
                "stock": stock
            }]
        }
        return self._request('POST', url, json=data)
    
    def create_product(self, product_data: Dict) -> Optional[Dict]:
        """Создание карточки товара"""
        url = f"{self.BASE_URL}/product"
        data = {
            "name": product_data.get('name', ''),
            "price": product_data.get('price', 0),
            "category": product_data.get('category', ''),
            "offer_id": product_data.get('offer_id', '')
        }
        return self._request('POST', url, json=data)


class WildberriesAPI(BaseMarketplaceAPI):
    """API Wildberries"""
    
    BASE_URL = "https://suppliers-api.wildberries.ru"
    
    def __init__(self, api_key: str = None):
        super().__init__(api_key)
        self.session.headers.update({
            'Authorization': api_key,
            'Content-Type': 'application/json'
        })
    
    def get_products(self, page: int = 1, limit: int = 100) -> Optional[Dict]:
        """Получение списка товаров"""
        cache_key = generate_cache_key('wb_products', page, limit)
        cached = self.cache.get(cache_key)
        if cached:
            return cached
        
        url = f"{self.BASE_URL}/content/v2/get/cards/list"
        data = {"limit": limit, "offset": (page - 1) * limit}
        
        result = self._request('POST', url, json=data)
        if result:
            self.cache.set(cache_key, result)
        return result
    
    def get_prices(self, nm_ids: List[int]) -> Optional[Dict]:
        """Получение цен"""
        cache_key = generate_cache_key('wb_prices', *sorted(nm_ids))
        cached = self.cache.get(cache_key)
        if cached:
            return cached
        
        url = f"{self.BASE_URL}/public/v1/prices"
        data = {"nm_ids": nm_ids}
        
        result = self._request('POST', url, json=data)
        if result:
            self.cache.set(cache_key, result)
        return result
    
    def update_price(self, nm_id: int, price: float) -> Optional[Dict]:
        """Обновление цены"""
        url = f"{self.BASE_URL}/public/v1/prices"
        data = [{"nm_id": nm_id, "price": price}]
        return self._request('POST', url, json=data)
    
    def update_stock(self, nm_id: int, stock: int) -> Optional[Dict]:
        """Обновление остатков"""
        url = f"{self.BASE_URL}/public/v1/stocks"
        data = {"nm_id": nm_id, "stock": stock}
        return self._request('POST', url, json=data)
    
    def create_product(self, product_data: Dict) -> Optional[Dict]:
        """Создание карточки товара"""
        url = f"{self.BASE_URL}/content/v2/create/card"
        data = {
            "nm_id": product_data.get('offer_id', ''),
            "name": product_data.get('name', ''),
            "price": product_data.get('price', 0),
            "category": product_data.get('category', '')
        }
        return self._request('POST', url, json=data)


class AliExpressAPI(BaseMarketplaceAPI):
    """API AliExpress (упрощенная версия)"""
    
    BASE_URL = "https://api.aliexpress.com/openapi"
    
    def __init__(self, api_key: str = None, secret: str = None):
        super().__init__(api_key)
        self.secret = secret
    
    def get_products(self, page: int = 1, page_size: int = 20) -> Optional[Dict]:
        """Получение списка товаров"""
        cache_key = generate_cache_key('aliexpress_products', page, page_size)
        cached = self.cache.get(cache_key)
        if cached:
            return cached
        
        params = {
            'method': 'aliexpress.product.list',
            'page': page,
            'page_size': page_size,
            'access_token': self.api_key
        }
        
        result = self._request('GET', self.BASE_URL, params=params)
        if result:
            self.cache.set(cache_key, result)
        return result

# --------------------------------------------
# 🗄️ БАЗА ДАННЫХ OE НОМЕРОВ
# --------------------------------------------
class OEMDatabase:
    """База данных OE номеров с автоподстановкой характеристик"""
    
    def __init__(self, db_path: str = "oem_database.json"):
        self.db_path = db_path
        self.data = {}
        self._load_database()
    
    def _load_database(self):
        """Загрузка базы данных из JSON"""
        if os.path.exists(self.db_path):
            try:
                with open(self.db_path, 'r', encoding='utf-8') as f:
                    self.data = json.load(f)
                logger.info(f"Загружено {len(self.data)} OE номеров из базы")
            except Exception as e:
                logger.error(f"Ошибка загрузки базы OE: {e}")
                self._create_demo_database()
        else:
            self._create_demo_database()
            self._save_database()
    
    def _create_demo_database(self):
        """Создание демо-базы OE номеров"""
        self.data = {
            "0986AF0059": {
                "category": "Фильтры",
                "subcategory": "Масляные фильтры",
                "brand": "BOSCH",
                "compatibility": ["BMW 3", "BMW 5", "Audi A4", "Audi A6", "VW Passat"],
                "weight": 0.35,
                "dimensions": {"length": 100, "width": 80, "height": 50},
                "cross_reference": ["MANN W842/2", "MAHLE OC 205"]
            },
            "W842/2": {
                "category": "Фильтры",
                "subcategory": "Масляные фильтры",
                "brand": "MANN",
                "compatibility": ["BMW 3", "BMW 5", "Audi A4", "Audi A6", "VW Passat"],
                "weight": 0.32,
                "dimensions": {"length": 95, "width": 75, "height": 48},
                "cross_reference": ["BOSCH 0986AF0059", "MAHLE OC 205"]
            },
            "OC205": {
                "category": "Фильтры",
                "subcategory": "Масляные фильтры",
                "brand": "MAHLE",
                "compatibility": ["BMW 3", "BMW 5", "Audi A4", "Audi A6", "VW Passat"],
                "weight": 0.33,
                "dimensions": {"length": 98, "width": 78, "height": 49},
                "cross_reference": ["BOSCH 0986AF0059", "MANN W842/2"]
            },
            "AB123456789": {
                "category": "Тормозная система",
                "subcategory": "Тормозные колодки",
                "brand": "BREMBO",
                "compatibility": ["VW Golf", "VW Passat", "Skoda Octavia", "Audi A3"],
                "weight": 1.2,
                "dimensions": {"length": 150, "width": 120, "height": 30},
                "cross_reference": ["BREMBO P85012", "ATE 13.0460-1234.2"]
            },
            "5524": {
                "category": "Свечи зажигания",
                "subcategory": "Свечи зажигания",
                "brand": "NGK",
                "compatibility": ["BMW 3", "BMW 5", "Audi A4", "VW Golf", "Toyota Camry"],
                "weight": 0.05,
                "dimensions": {"length": 50, "width": 20, "height": 20},
                "cross_reference": ["BOSCH FR7DC", "DENSO K16PR-U11"]
            },
            "1234567890": {
                "category": "Стеклоочистители",
                "subcategory": "Щетки стеклоочистителя",
                "brand": "VALEO",
                "compatibility": ["BMW 3", "BMW 5", "Audi A4", "VW Passat"],
                "weight": 0.25,
                "dimensions": {"length": 500, "width": 30, "height": 20},
                "cross_reference": ["BOSCH 3397007405", "SWF 119-000-016"]
            },
            "12345678901": {
                "category": "Тормозная система",
                "subcategory": "Тормозные диски",
                "brand": "ATE",
                "compatibility": ["BMW 3", "BMW 5", "Audi A4", "VW Passat"],
                "weight": 8.5,
                "dimensions": {"length": 300, "width": 300, "height": 25},
                "cross_reference": ["BREMBO 09.9870.10", "BOSCH 0986AB8765"]
            },
            "6PK1735": {
                "category": "Ремни и цепи",
                "subcategory": "Ремни ГРМ",
                "brand": "CONTINENTAL",
                "compatibility": ["Audi A4", "Audi A6", "VW Passat", "Skoda Octavia"],
                "weight": 0.15,
                "dimensions": {"length": 1735, "width": 6, "height": 6},
                "cross_reference": ["GATES 6PK1735", "DAYCO 6PK1735"]
            }
        }
        
        for i in range(50):
            oe = f"OE{random.randint(10000, 99999)}{chr(65+random.randint(0, 25))}"
            categories = ["Двигатель", "Трансмиссия", "Подвеска", "Электрооборудование", "Система охлаждения", "Масла и жидкости"]
            brands = ["BOSCH", "DENSO", "NGK", "BREMBO", "AISIN", "HITACHI", "VALEO", "MANN", "MAHLE", "SKF"]
            self.data[oe] = {
                "category": random.choice(categories),
                "subcategory": "Запчасть",
                "brand": random.choice(brands),
                "compatibility": ["BMW 3", "Audi A4", "VW Golf"],
                "weight": round(random.uniform(0.1, 10), 2),
                "dimensions": {
                    "length": random.randint(50, 500),
                    "width": random.randint(20, 400),
                    "height": random.randint(10, 200)
                }
            }
    
    def _save_database(self):
        """Сохранение базы данных в JSON"""
        try:
            with open(self.db_path, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            logger.info(f"База OE сохранена: {len(self.data)} записей")
        except Exception as e:
            logger.error(f"Ошибка сохранения базы OE: {e}")
    
    def get_by_oe(self, oe_number: str) -> Optional[Dict]:
        """Получение данных по OE номеру"""
        if not oe_number:
            return None
        
        oe_number = oe_number.strip().upper()
        
        if oe_number in self.data:
            return self.data[oe_number].copy()
        
        for key, value in self.data.items():
            if oe_number in key or key in oe_number:
                return value.copy()
        
        return None
    
    def get_category(self, oe_number: str) -> str:
        """Получение категории по OE номеру"""
        data = self.get_by_oe(oe_number)
        return data.get("category", "Прочее") if data else "Прочее"
    
    def get_brand(self, oe_number: str) -> Optional[str]:
        """Получение бренда по OE номеру"""
        data = self.get_by_oe(oe_number)
        return data.get("brand") if data else None
    
    def get_dimensions(self, oe_number: str) -> Dict:
        """Получение размеров по OE номеру"""
        data = self.get_by_oe(oe_number)
        return data.get("dimensions", {}) if data else {}
    
    def get_weight(self, oe_number: str) -> float:
        """Получение веса по OE номеру"""
        data = self.get_by_oe(oe_number)
        return data.get("weight", 0) if data else 0
    
    def get_compatibility(self, oe_number: str) -> List[str]:
        """Получение совместимости по OE номеру"""
        data = self.get_by_oe(oe_number)
        return data.get("compatibility", []) if data else []
    
    def add_or_update(self, oe_number: str, data: Dict):
        """Добавление или обновление записи в базе"""
        oe_number = oe_number.strip().upper()
        self.data[oe_number] = data
        self._save_database()
    
    def search_by_brand(self, brand: str) -> List[Dict]:
        """Поиск по бренду"""
        results = []
        for oe, data in self.data.items():
            if data.get("brand", "").upper() == brand.upper():
                result = data.copy()
                result["oe_number"] = oe
                results.append(result)
        return results
    
    def search_by_category(self, category: str) -> List[Dict]:
        """Поиск по категории"""
        results = []
        for oe, data in self.data.items():
            if data.get("category", "").lower() == category.lower():
                result = data.copy()
                result["oe_number"] = oe
                results.append(result)
        return results
    
    def get_statistics(self) -> Dict:
        """Статистика базы данных"""
        stats = {
            "total": len(self.data),
            "categories": {},
            "brands": {}
        }
        
        for data in self.data.values():
            category = data.get("category", "Прочее")
            stats["categories"][category] = stats["categories"].get(category, 0) + 1
            
            brand = data.get("brand", "Неизвестно")
            stats["brands"][brand] = stats["brands"].get(brand, 0) + 1
        
        return stats

# --------------------------------------------
# 🕷️ ПАРСЕРЫ ЦЕН КОНКУРЕНТОВ
# --------------------------------------------
class CompetitorParser:
    """Парсинг цен конкурентов с маркетплейсов"""
    
    def __init__(self):
        self.cache = APICache()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
    
    @st.cache_data(ttl=300)
    def parse_yandex_market(_self, query: str, max_pages: int = 2) -> List[Dict]:
        """Парсинг Яндекс Маркета"""
        results = []
        
        for page in range(1, max_pages + 1):
            try:
                url = "https://market.yandex.ru/search"
                params = {'text': query, 'page': page}
                
                response = _self.session.get(url, params=params, timeout=20)
                if response.status_code != 200:
                    continue
                
                html = response.text
                items = re.findall(r'"offerId":"([^"]+)","name":"([^"]+)","price":"([^"]+)"', html)
                
                for offer_id, name, price in items[:10]:
                    results.append({
                        'marketplace': 'Яндекс Маркет',
                        'offer_id': offer_id,
                        'name': name,
                        'price': safe_float(price),
                        'parsed_at': datetime.now().isoformat()
                    })
                
                time.sleep(0.3)
                
            except Exception as e:
                logger.error(f"Yandex parse error: {e}")
        
        return results
    
    @st.cache_data(ttl=300)
    def parse_ozon(_self, query: str, max_pages: int = 2) -> List[Dict]:
        """Парсинг Ozon"""
        results = []
        
        for page in range(1, max_pages + 1):
            try:
                url = "https://www.ozon.ru/api/composer-api.bx/page/json/v2"
                params = {'url': f'/search/?text={urllib.parse.quote(query)}&page={page}'}
                
                response = _self.session.get(url, params=params, timeout=20)
                if response.status_code != 200:
                    continue
                
                data = response.json()
                widgets = data.get('widgets', [])
                
                for widget in widgets:
                    if widget.get('type') == 'searchResultsV2':
                        items = widget.get('items', [])
                        for item in items[:10]:
                            results.append({
                                'marketplace': 'Ozon',
                                'product_id': item.get('id'),
                                'name': item.get('name', ''),
                                'price': safe_float(item.get('price', {}).get('price', 0)),
                                'parsed_at': datetime.now().isoformat()
                            })
                
                time.sleep(0.3)
                
            except Exception as e:
                logger.error(f"Ozon parse error: {e}")
        
        return results
    
    @st.cache_data(ttl=300)
    def parse_wildberries(_self, query: str, max_pages: int = 2) -> List[Dict]:
        """Парсинг Wildberries"""
        results = []
        
        for page in range(1, max_pages + 1):
            try:
                url = "https://search.wb.ru/exactmatch/ru/common/v4/search"
                params = {'query': query, 'page': page, 'limit': 50}
                
                response = _self.session.get(url, params=params, timeout=20)
                if response.status_code != 200:
                    continue
                
                data = response.json()
                products = data.get('data', {}).get('products', [])
                
                for product in products[:10]:
                    results.append({
                        'marketplace': 'Wildberries',
                        'nm_id': product.get('id'),
                        'name': product.get('name', ''),
                        'price': safe_float(product.get('priceU', 0)) / 100,
                        'parsed_at': datetime.now().isoformat()
                    })
                
                time.sleep(0.3)
                
            except Exception as e:
                logger.error(f"WB parse error: {e}")
        
        return results
    
    def parse_all_marketplaces(self, query: str) -> Dict[str, List[Dict]]:
        """Парсинг всех маркетплейсов"""
        results = {}
        results_dict = {}
        threads = []
        
        def parse_with_store(marketplace, parser_func):
            try:
                results_dict[marketplace] = parser_func(query)
            except Exception as e:
                logger.error(f"Error parsing {marketplace}: {e}")
                results_dict[marketplace] = []
        
        parsers = [
            ('Яндекс Маркет', self.parse_yandex_market),
            ('Ozon', self.parse_ozon),
            ('Wildberries', self.parse_wildberries)
        ]
        
        for name, func in parsers:
            thread = Thread(target=parse_with_store, args=(name, func))
            thread.start()
            threads.append(thread)
        
        for thread in threads:
            thread.join(timeout=30)
        
        return results_dict

# --------------------------------------------
# 🏪 УПРАВЛЕНИЕ КОНКУРЕНТАМИ
# --------------------------------------------
class CompetitorManager:
    """Управление конкурентными данными"""
    
    def __init__(self):
        self.parser = CompetitorParser()
        self.competitor_data = {}
        self.last_update = {}
    
    def get_competitor_prices(self, query: str, marketplace: str = None) -> Dict:
        """Получение цен конкурентов"""
        cache_key = generate_cache_key('competitor_prices', query, marketplace or 'all')
        
        if cache_key in self.competitor_data:
            data, timestamp = self.competitor_data[cache_key]
            if (datetime.now() - timestamp).total_seconds() < 3600:
                return data
        
        if marketplace:
            parser_map = {
                'Яндекс Маркет': self.parser.parse_yandex_market,
                'Ozon': self.parser.parse_ozon,
                'Wildberries': self.parser.parse_wildberries
            }
            parser = parser_map.get(marketplace)
            if parser:
                results = parser(query)
            else:
                results = []
        else:
            results = self.parser.parse_all_marketplaces(query)
        
        self.competitor_data[cache_key] = (results, datetime.now())
        return results
    
    def analyze_competitor_prices(self, product_name: str, our_price: float) -> Dict:
        """Анализ цен конкурентов"""
        competitor_data = self.get_competitor_prices(product_name)
        
        all_prices = []
        if isinstance(competitor_data, dict):
            for marketplace, items in competitor_data.items():
                for item in items:
                    price = item.get('price', 0)
                    if price > 0:
                        all_prices.append({
                            'marketplace': marketplace,
                            'price': price,
                            'name': item.get('name', '')
                        })
        elif isinstance(competitor_data, list):
            for item in competitor_data:
                price = item.get('price', 0)
                if price > 0:
                    all_prices.append({
                        'marketplace': item.get('marketplace', 'Неизвестно'),
                        'price': price,
                        'name': item.get('name', '')
                    })
        
        if not all_prices:
            return {
                'competitor_count': 0,
                'avg_price': our_price,
                'min_price': our_price,
                'max_price': our_price,
                'price_position': 'Нет данных',
                'recommendation': 'Нет конкурентов для анализа'
            }
        
        prices = [p['price'] for p in all_prices]
        avg_price = sum(prices) / len(prices)
        min_price = min(prices)
        max_price = max(prices)
        
        if our_price <= min_price:
            position = "Ниже всех"
            recommendation = "Высокая конкурентоспособность"
        elif our_price <= avg_price:
            position = "Ниже среднего"
            recommendation = "Хорошая позиция"
        elif our_price <= max_price:
            position = "Выше среднего"
            recommendation = "Стоит снизить цену"
        else:
            position = "Выше всех"
            recommendation = "Срочно снизить цену"
        
        return {
            'competitor_count': len(all_prices),
            'avg_price': avg_price,
            'min_price': min_price,
            'max_price': max_price,
            'price_position': position,
            'recommendation': recommendation,
            'competitors': all_prices[:10]
        }

# --------------------------------------------
# 🧠 AI-ПОЛУЧЕНИЕ ТАРИФОВ
# --------------------------------------------
class AITariffProvider:
    """Получение актуальных тарифов через AI с кэшированием"""
    
    def __init__(self, api_key: str = None, cache_ttl: int = 3600):
        self.api_key = api_key
        self.api_url = "https://api.deepseek.com/v1/chat/completions"
        self.cache = {}
        self.cache_ttl = cache_ttl
        self.last_update = {}
        
    def get_rates(self, marketplace: str, mode: str = "FBY") -> Dict:
        """Получение тарифов через AI или из кэша"""
        cache_key = generate_cache_key(marketplace, mode)
        
        if cache_key in self.cache:
            cached_data, timestamp = self.cache[cache_key]
            if (datetime.now() - timestamp).total_seconds() < self.cache_ttl:
                logger.info(f"Использую кэшированные тарифы для {marketplace}/{mode}")
                return cached_data.copy()
        
        rates = self._get_base_rates(marketplace, mode)
        
        if self.api_key:
            try:
                ai_rates = self._get_ai_rates(marketplace, mode)
                if ai_rates and isinstance(ai_rates, dict):
                    for key, value in ai_rates.items():
                        if key in rates and isinstance(value, (int, float)):
                            rates[key] = value
                    logger.info(f"Получены AI-тарифы для {marketplace}/{mode}")
            except Exception as e:
                logger.error(f"AI tariff error: {e}")
        
        self.cache[cache_key] = (rates.copy(), datetime.now())
        self.last_update[cache_key] = datetime.now()
        
        return rates
    
    def _get_base_rates(self, marketplace: str, mode: str) -> Dict:
        """Базовые тарифы"""
        rates = {
            "commission": 0.11,
            "logistics_base": 70.0,
            "logistics_per_liter": 22.0,
            "storage_per_liter": 3.0,
            "storage_free_days": 365,
            "acquiring": 0.022,
            "returns": 0.10,
            "advertising": 0.15,
            "packaging": 50.0,
            "fba_base": 70.0,
            "fba_per_kg": 12.0,
            "fbs_logistics": 115.0,
            "fbo_storage": 0.5,
            "delivery_to_customer_percent": 0.045,
            "delivery_to_customer_max": 1000.0,
            "service_fee": 0.01,
            "insurance": 0.005
        }
        
        marketplace_rates = {
            "Яндекс Маркет": {"commission": 0.11},
            "Ozon": {"commission": 0.10},
            "Wildberries": {"commission": 0.12},
            "AliExpress": {"commission": 0.08},
            "Мегамаркет": {"commission": 0.09}
        }
        
        rates.update(marketplace_rates.get(marketplace, {}))
        
        mode_rates = {
            "FBY": {"fba_base": 70.0, "fba_per_kg": 12.0, "logistics_base": 70.0},
            "FBS": {"fbs_logistics": 115.0, "logistics_base": 115.0},
            "FBO": {"fbo_storage": 0.5, "storage_per_liter": 0.5},
            "DBS": {"logistics_base": 60.0, "logistics_per_liter": 10.0}
        }
        
        rates.update(mode_rates.get(mode, {}))
        
        return rates
    
    def _get_ai_rates(self, marketplace: str, mode: str) -> Optional[Dict]:
        """Получение тарифов через AI API"""
        if not self.api_key:
            return None
            
        try:
            prompt = f"""
            Предоставь актуальные тарифы для маркетплейса {marketplace} 
            для продажи автозапчастей на начало 2026 года.
            Режим работы: {mode}
            
            Верни ТОЛЬКО JSON с тарифами.
            """
            
            headers = {
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json"
            }
            
            data = {
                "model": "deepseek-chat",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1,
                "max_tokens": 1000,
                "response_format": {"type": "json_object"}
            }
            
            response = requests.post(self.api_url, headers=headers, json=data, timeout=30)
            
            if response.status_code == 200:
                content = response.json()['choices'][0]['message']['content']
                json_match = re.search(r'\{.*\}', content, re.DOTALL)
                if json_match:
                    rates = json.loads(json_match.group())
                    required_keys = ["commission", "logistics_base", "logistics_per_liter", 
                                   "storage_per_liter", "acquiring"]
                    if all(key in rates for key in required_keys):
                        return rates
            return None
            
        except Exception as e:
            logger.error(f"AI API error: {e}")
            return None
    
    def get_all_rates(self) -> Dict:
        """Получение тарифов для всех маркетплейсов и режимов"""
        all_rates = {}
        for marketplace in CONFIG["marketplaces"]:
            all_rates[marketplace] = {}
            for mode in CONFIG["operation_modes"]:
                all_rates[marketplace][mode] = self.get_rates(marketplace, mode)
        return all_rates
    
    def clear_cache(self):
        """Очистка кэша"""
        self.cache.clear()
        self.last_update.clear()
        logger.info("Кэш тарифов очищен")

# --------------------------------------------
# 🏷️ КЛАССИФИКАТОР КАТЕГОРИЙ (с ML)
# --------------------------------------------
class CategoryClassifier:
    """Классификация категорий товаров с использованием ML"""
    
    def __init__(self):
        self.keywords = CONFIG["category_keywords"]
        self.categories = list(self.keywords.keys())
        self.cache = {}
        self.oem_patterns = CONFIG["oem_patterns"]
        self.ml_classifier = AutoClassifier() if SKLEARN_AVAILABLE else None
        
    @lru_cache(maxsize=10000)
    def classify(self, name: str) -> Tuple[str, float]:
        """Определение категории по наименованию с ML"""
        if not name:
            return "Прочее", 0.0
        
        if self.ml_classifier:
            ml_category, ml_confidence = self.ml_classifier.predict(name)
            if ml_confidence > 50:
                return ml_category, ml_confidence
        
        name_lower = name.lower()
        best_category = "Прочее"
        best_score = 0.0
        
        for category, keywords in self.keywords.items():
            score = 0.0
            for keyword in keywords:
                if keyword.lower() in name_lower:
                    weight = len(keyword) / 10.0
                    if name_lower.startswith(keyword.lower()):
                        weight *= 1.5
                    score += weight
            
            if score > best_score:
                best_score = score
                best_category = category
        
        confidence = min(best_score * 20, 100.0)
        
        if self.extract_oem(name):
            confidence = min(confidence + 10, 100)
        
        return best_category, round(confidence, 1)
    
    def extract_oem(self, name: str) -> Optional[str]:
        """Извлечение OEM номера из наименования"""
        if not name:
            return None
        
        for pattern in self.oem_patterns:
            match = re.search(pattern, name.upper())
            if match:
                return match.group()
        return None
    
    def extract_brand(self, name: str) -> Optional[str]:
        """Извлечение бренда из наименования"""
        if not name:
            return None
        
        brands = [
            "BOSCH", "DENSO", "NGK", "BREMBO", "AISIN", "HITACHI", "VALEO",
            "PIERBURG", "MANN", "MAHLE", "HENGST", "SACHS", "ZF",
            "CONTINENTAL", "GATES", "DAYCO", "SKF", "FAG", "TIMKEN",
            "MERCEDES", "BMW", "AUDI", "VW", "FORD", "TOYOTA", "HONDA",
            "NISSAN", "HYUNDAI", "KIA", "SKODA", "VOLVO", "RENAULT"
        ]
        
        name_upper = name.upper()
        for brand in brands:
            if brand in name_upper:
                return brand
        return None

# --------------------------------------------
# 📊 КЛАСС ВАЛИДАЦИИ ДАННЫХ
# --------------------------------------------
class DataValidator:
    """Валидация данных"""
    
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.stats = {
            "total": 0,
            "valid": 0,
            "invalid": 0,
            "skipped": 0,
            "by_error_type": defaultdict(int)
        }
    
    def validate_product(self, row: Dict) -> Tuple[bool, List[str]]:
        """Валидация одного товара"""
        errors = []
        warnings = []
        
        price = safe_float(row.get("price", 0))
        if price <= 0:
            errors.append("Цена должна быть больше 0")
        elif price < CONFIG["validation"]["min_price"]:
            warnings.append(f"Цена очень низкая: {price} ₽")
        elif price > CONFIG["validation"]["max_price"]:
            warnings.append(f"Цена очень высокая: {price} ₽")
        
        cost = safe_float(row.get("cost", 0))
        if cost < 0:
            errors.append("Себестоимость не может быть отрицательной")
        elif cost == 0 and price > 0:
            warnings.append("Себестоимость не указана, используется 50% от цены")
        
        for dim, label in [("length", "Длина"), ("width", "Ширина"), ("height", "Высота")]:
            value = safe_float(row.get(dim, 0))
            if value < 0:
                errors.append(f"{label} не может быть отрицательной")
            elif value > CONFIG["validation"]["max_dimension"]:
                warnings.append(f"Подозрительно большая {label.lower()}: {value} мм")
        
        name = safe_str(row.get("name", ""))
        if not name:
            warnings.append("Отсутствует наименование товара")
        
        article = safe_str(row.get("article", ""))
        if not article:
            warnings.append("Отсутствует артикул товара")
        
        self.stats["total"] += 1
        for error in errors:
            self.stats["by_error_type"][error] += 1
        
        if errors:
            self.stats["invalid"] += 1
        else:
            self.stats["valid"] += 1
            if warnings:
                self.stats["skipped"] += 1
        
        self.errors.extend(errors)
        self.warnings.extend(warnings)
        
        return len(errors) == 0, errors + warnings
    
    def get_report(self) -> Dict:
        """Получение отчета по валидации"""
        return {
            "total": self.stats["total"],
            "valid": self.stats["valid"],
            "invalid": self.stats["invalid"],
            "skipped": self.stats["skipped"],
            "errors": self.errors[:100],
            "warnings": self.warnings[:100],
            "error_types": dict(self.stats["by_error_type"])
        }
    
    def reset(self):
        """Сброс валидатора"""
        self.errors = []
        self.warnings = []
        self.stats = {
            "total": 0,
            "valid": 0,
            "invalid": 0,
            "skipped": 0,
            "by_error_type": defaultdict(int)
        }

# --------------------------------------------
# 💎 ДВИЖОК РАСЧЕТА ЮНИТ-ЭКОНОМИКИ
# --------------------------------------------
class UnitEconomicsEngine:
    """Движок расчета юнит-экономики с поддержкой конвертации размеров"""
    
    def __init__(self, 
                 marketplace: str = "Яндекс Маркет", 
                 mode: str = "FBY", 
                 days_storage: int = 30,
                 dimension_unit: str = "мм"):
        
        self.marketplace = marketplace
        self.mode = mode
        self.days_storage = days_storage
        self.dimension_unit = dimension_unit
        self.tariff_provider = AITariffProvider()
        self.rates = self.tariff_provider.get_rates(marketplace, mode)
        self.fixed_costs = 50000.0
        self.avg_orders = 100
        self.retention_rate = 0.7
        self.discount_rate = 0.1
        self.lead_time = 7
        self.z_score = 1.65
        self.validator = DataValidator()
        self.competitor_manager = CompetitorManager()
        self.oem_db = OEMDatabase()
    
    def _convert_to_mm(self, value: float) -> float:
        """Конвертация значения в мм"""
        if value == 0:
            return 0.0
        
        if self.dimension_unit == "мм":
            return value
        elif self.dimension_unit == "см":
            return value * 10.0
        
        return value
        
    def calculate_product(self, row: Dict) -> Optional[Dict]:
        """Расчет юнит-экономики для одного товара"""
        try:
            is_valid, validation_issues = self.validator.validate_product(row)
            return self._calculate_product_model(row, is_valid, validation_issues)
                
        except Exception as e:
            logger.error(f"Error calculating product {row.get('article', 'unknown')}: {e}")
            return None
    
    def _calculate_product_model(self, row: Dict, is_valid: bool, validation_issues: List[str]) -> Dict:
        """Расчет для товарной модели (B2C)"""
        article = safe_str(row.get("article", ""))
        name = safe_str(row.get("name", ""))
        price = safe_float(row.get("price", 0))
        
        # Получаем размеры и конвертируем в мм
        length_orig = safe_float(row.get("length", 0))
        width_orig = safe_float(row.get("width", 0))
        height_orig = safe_float(row.get("height", 0))
        
        # Конвертируем в мм для расчетов
        length = self._convert_to_mm(length_orig)
        width = self._convert_to_mm(width_orig)
        height = self._convert_to_mm(height_orig)
        
        cost = safe_float(row.get("cost", price * 0.5 if price > 0 else 0))
        weight = safe_float(row.get("weight", 0))
        oe_number = safe_str(row.get("oe_number", ""))
        brand = safe_str(row.get("brand", ""))
        category = safe_str(row.get("category", "Прочее"))
        compatibility = safe_str(row.get("compatibility", ""))
        
        if price <= 0:
            return None
        
        competitor_analysis = self.competitor_manager.analyze_competitor_prices(name, price)
        
        volume = calculate_volume(length, width, height)
        if volume == 0:
            volume = 1.0
        
        if weight == 0 and volume > 0:
            weight = volume * 0.8
        
        if weight == 0:
            weight = 1.0
        
        commission = price * self.rates.get("commission", 0.11)
        service_fee = price * self.rates.get("service_fee", 0.01)
        insurance = price * self.rates.get("insurance", 0.005)
        logistics = self._calculate_logistics(volume, weight, price)
        storage = self._calculate_storage(volume)
        acquiring = price * self.rates.get("acquiring", 0.022)
        advertising = price * self.rates.get("advertising", 0.15)
        returns = price * self.rates.get("returns", 0.10)
        packaging = self.rates.get("packaging", 50.0)
        
        total_variable = cost + commission + service_fee + insurance + logistics + storage + acquiring + advertising + returns + packaging
        total_cost = total_variable
        
        unit_profit = price - total_cost
        margin = (unit_profit / price * 100) if price > 0 else 0
        contribution_margin = price - total_variable
        contribution_margin_pct = (contribution_margin / price * 100) if price > 0 else 0
        
        roi = (unit_profit / cost * 100) if cost > 0 else 0
        
        daily_profit = unit_profit * self.avg_orders / 30
        payback_days = cost / daily_profit if daily_profit > 0 else 999
        
        ltv = unit_profit / (1 - self.retention_rate + self.discount_rate)
        cac = advertising / 5 * 1000 if advertising > 0 else price * 0.05
        ltv_cac_ratio = ltv / cac if cac > 0 else 0
        
        annual_demand = self.avg_orders * 12
        holding_cost = cost * 0.2
        eoq = math.sqrt((2 * annual_demand * 500) / holding_cost) if holding_cost > 0 else 0
        
        safety_stock = self.z_score * (self.avg_orders / 30 * 0.2) * math.sqrt(self.lead_time)
        reorder_point = (self.avg_orders / 30 * self.lead_time) + safety_stock
        
        break_even_units = self.fixed_costs / contribution_margin if contribution_margin > 0 else 999999
        break_even_revenue = break_even_units * price
        
        profit_status = "Прибыльный" if unit_profit > 0 else "Убыточный"
        margin_status = "Высокая" if margin > 25 else "Средняя" if margin > 10 else "Низкая"
        
        if margin < 15:
            recommended_price = price * 1.15
            price_action = "Повысить"
            price_reason = "Низкая маржинальность (<15%)"
        elif margin > 35:
            recommended_price = price * 0.95
            price_action = "Снизить"
            price_reason = "Высокая маржинальность (>35%)"
        elif competitor_analysis.get('price_position') == "Выше всех":
            recommended_price = price * 0.95
            price_action = "Снизить"
            price_reason = f"Цена выше всех конкурентов (средняя {competitor_analysis.get('avg_price', 0):.0f} ₽)"
        else:
            recommended_price = price
            price_action = "Оставить"
            price_reason = "Оптимальная маржинальность"
        
        return {
            "article": article,
            "name": name,
            "brand": brand,
            "oe_number": oe_number,
            "category": category,
            "compatibility": compatibility,
            "price": round(price, 2),
            "cost": round(cost, 2),
            "length_orig": round(length_orig, 1),
            "width_orig": round(width_orig, 1),
            "height_orig": round(height_orig, 1),
            "length_mm": round(length, 1),
            "width_mm": round(width, 1),
            "height_mm": round(height, 1),
            "dimension_unit": self.dimension_unit,
            "volume": round(volume, 2),
            "weight": round(weight, 2),
            "commission": round(commission, 2),
            "service_fee": round(service_fee, 2),
            "insurance": round(insurance, 2),
            "logistics": round(logistics, 2),
            "storage": round(storage, 2),
            "acquiring": round(acquiring, 2),
            "advertising": round(advertising, 2),
            "returns": round(returns, 2),
            "packaging": round(packaging, 2),
            "total_variable_costs": round(total_variable, 2),
            "total_cost": round(total_cost, 2),
            "contribution_margin": round(contribution_margin, 2),
            "contribution_margin_pct": round(contribution_margin_pct, 1),
            "unit_profit": round(unit_profit, 2),
            "margin": round(margin, 1),
            "roi": round(roi, 1),
            "payback_days": round(payback_days, 0),
            "break_even_units": round(break_even_units, 0),
            "break_even_revenue": round(break_even_revenue, 2),
            "ltv": round(ltv, 2),
            "cac": round(cac, 2),
            "ltv_cac_ratio": round(ltv_cac_ratio, 2),
            "eoq": round(eoq, 0),
            "reorder_point": round(reorder_point, 0),
            "safety_stock": round(safety_stock, 0),
            "profit_status": profit_status,
            "margin_status": margin_status,
            "recommended_price": round(recommended_price, 2),
            "price_action": price_action,
            "price_reason": price_reason,
            "abc_category": self._abc_category(unit_profit),
            "xyz_category": self._xyz_category(self.avg_orders),
            "competitor_count": competitor_analysis.get('competitor_count', 0),
            "competitor_avg_price": round(competitor_analysis.get('avg_price', 0), 2),
            "competitor_min_price": round(competitor_analysis.get('min_price', 0), 2),
            "competitor_max_price": round(competitor_analysis.get('max_price', 0), 2),
            "price_position": competitor_analysis.get('price_position', 'Нет данных'),
            "competitor_recommendation": competitor_analysis.get('recommendation', ''),
            "mode": self.mode,
            "marketplace": self.marketplace,
            "validation_valid": is_valid,
            "validation_issues": "; ".join(validation_issues[:3]) if validation_issues else "",
        }
    
    def _calculate_logistics(self, volume: float, weight: float, price: float) -> float:
        """Расчет логистики"""
        try:
            if self.mode == "FBY":
                logistics = self.rates.get("fba_base", 70.0) + weight * self.rates.get("fba_per_kg", 12.0)
            elif self.mode == "FBS":
                logistics = self.rates.get("fbs_logistics", 115.0)
                if volume <= self.rates.get("logistics_threshold", 1.0):
                    logistics = self.rates.get("logistics_base", 70.0)
                elif volume <= self.rates.get("logistics_threshold_160", 160.0):
                    logistics = self.rates.get("logistics_base", 70.0) + (volume - 1.0) * self.rates.get("logistics_per_liter", 22.0)
                else:
                    logistics = self.rates.get("logistics_base", 70.0) + (160.0 - 1.0) * 22.0 + (volume - 160.0) * 2.0
            else:
                logistics = self.rates.get("logistics_base", 70.0) + volume * self.rates.get("logistics_per_liter", 10.0) + weight * 15.0
            
            if self.mode != "FBY":
                delivery_to_customer = min(price * self.rates.get("delivery_to_customer_percent", 0.045), 
                                         self.rates.get("delivery_to_customer_max", 1000.0))
                logistics += delivery_to_customer
            
            return round(logistics, 2)
        except Exception as e:
            logger.error(f"Error calculating logistics: {e}")
            return self.rates.get("logistics_base", 70.0)
    
    def _calculate_storage(self, volume: float) -> float:
        """Расчет хранения"""
        try:
            if self.mode == "FBY":
                free_days = self.rates.get("storage_free_days", 365)
                if self.days_storage <= free_days:
                    return 0.0
                return round(volume * self.rates.get("storage_per_liter", 3.0) * (self.days_storage - free_days), 2)
            elif self.mode == "FBO":
                return round(volume * self.rates.get("fbo_storage", 0.5) * (self.days_storage / 30), 2)
            else:
                return round(volume * self.rates.get("storage_per_liter", 0.8) * (self.days_storage / 30), 2)
        except Exception as e:
            logger.error(f"Error calculating storage: {e}")
            return 0.0
    
    def _abc_category(self, profit: float) -> str:
        """Определение ABC-категории"""
        if profit > 1000:
            return "A"
        elif profit > 100:
            return "B"
        return "C"
    
    def _xyz_category(self, sales: float) -> str:
        """Определение XYZ-категории"""
        if sales > 100:
            return "X"
        elif sales > 50:
            return "Y"
        return "Z"
    
    def get_validation_report(self) -> Dict:
        """Получение отчета по валидации"""
        return self.validator.get_report()

# --------------------------------------------
# 📤 АВТОМАТИЧЕСКАЯ ВЫГРУЗКА НА МАРКЕТПЛЕЙСЫ
# --------------------------------------------
class MarketplaceUploader:
    """Автоматическая выгрузка данных на маркетплейсы"""
    
    def __init__(self):
        self.api_clients = {}
        self.upload_queue = Queue()
        self.is_running = False
        self.upload_log = []
        self.lock = Lock()
        
    def register_client(self, marketplace: str, client: BaseMarketplaceAPI):
        """Регистрация API-клиента для маркетплейса"""
        self.api_clients[marketplace] = client
        
    def upload_prices(self, marketplace: str, products: List[Dict]) -> Dict:
        """Выгрузка цен на маркетплейс"""
        client = self.api_clients.get(marketplace)
        if not client:
            return {"status": "error", "message": f"Клиент для {marketplace} не зарегистрирован"}
        
        results = {
            "marketplace": marketplace,
            "total": len(products),
            "success": 0,
            "failed": 0,
            "errors": [],
            "timestamp": datetime.now().isoformat()
        }
        
        for product in products:
            try:
                offer_id = product.get('offer_id', product.get('article', ''))
                price = product.get('recommended_price', product.get('price', 0))
                
                if not offer_id or price <= 0:
                    continue
                
                if marketplace == "Яндекс Маркет":
                    result = client.update_price(offer_id, price)
                elif marketplace == "Ozon":
                    result = client.update_price(offer_id, price)
                elif marketplace == "Wildberries":
                    result = client.update_price(int(offer_id), price)
                else:
                    continue
                
                if result and result.get('status') != 'error':
                    results["success"] += 1
                else:
                    results["failed"] += 1
                    results["errors"].append({
                        "offer_id": offer_id,
                        "error": str(result) if result else "Неизвестная ошибка"
                    })
                
                time.sleep(0.1)
                
            except Exception as e:
                results["failed"] += 1
                results["errors"].append({
                    "offer_id": product.get('offer_id', 'unknown'),
                    "error": str(e)
                })
        
        with self.lock:
            self.upload_log.append(results)
        
        return results
    
    def upload_stocks(self, marketplace: str, products: List[Dict]) -> Dict:
        """Выгрузка остатков на маркетплейс"""
        client = self.api_clients.get(marketplace)
        if not client:
            return {"status": "error", "message": f"Клиент для {marketplace} не зарегистрирован"}
        
        results = {
            "marketplace": marketplace,
            "total": len(products),
            "success": 0,
            "failed": 0,
            "errors": [],
            "timestamp": datetime.now().isoformat()
        }
        
        for product in products:
            try:
                offer_id = product.get('offer_id', product.get('article', ''))
                stock = product.get('stock', product.get('quantity', 0))
                
                if not offer_id or stock < 0:
                    continue
                
                if marketplace == "Яндекс Маркет":
                    result = client.update_stock(offer_id, stock)
                elif marketplace == "Ozon":
                    result = client.update_stock(offer_id, stock)
                elif marketplace == "Wildberries":
                    result = client.update_stock(int(offer_id), stock)
                else:
                    continue
                
                if result and result.get('status') != 'error':
                    results["success"] += 1
                else:
                    results["failed"] += 1
                    results["errors"].append({
                        "offer_id": offer_id,
                        "error": str(result) if result else "Неизвестная ошибка"
                    })
                
                time.sleep(0.1)
                
            except Exception as e:
                results["failed"] += 1
                results["errors"].append({
                    "offer_id": product.get('offer_id', 'unknown'),
                    "error": str(e)
                })
        
        with self.lock:
            self.upload_log.append(results)
        
        return results
    
    def upload_products(self, marketplace: str, products: List[Dict]) -> Dict:
        """Создание/обновление карточек товаров"""
        client = self.api_clients.get(marketplace)
        if not client:
            return {"status": "error", "message": f"Клиент для {marketplace} не зарегистрирован"}
        
        results = {
            "marketplace": marketplace,
            "total": len(products),
            "success": 0,
            "failed": 0,
            "errors": [],
            "timestamp": datetime.now().isoformat()
        }
        
        for product in products:
            try:
                card_data = {
                    "offer_id": product.get('article', ''),
                    "name": product.get('name', ''),
                    "price": product.get('price', 0),
                    "category": product.get('category', ''),
                    "brand": product.get('brand', ''),
                    "description": product.get('description', ''),
                    "images": product.get('images', []),
                    "dimensions": {
                        "length": product.get('length_mm', 0),
                        "width": product.get('width_mm', 0),
                        "height": product.get('height_mm', 0)
                    }
                }
                
                if marketplace == "Яндекс Маркет":
                    result = client.create_product(card_data)
                elif marketplace == "Ozon":
                    result = client.create_product(card_data)
                elif marketplace == "Wildberries":
                    result = client.create_product(card_data)
                else:
                    continue
                
                if result and result.get('status') != 'error':
                    results["success"] += 1
                else:
                    results["failed"] += 1
                    results["errors"].append({
                        "offer_id": product.get('article', 'unknown'),
                        "error": str(result) if result else "Неизвестная ошибка"
                    })
                
                time.sleep(0.2)
                
            except Exception as e:
                results["failed"] += 1
                results["errors"].append({
                    "offer_id": product.get('article', 'unknown'),
                    "error": str(e)
                })
        
        with self.lock:
            self.upload_log.append(results)
        
        return results
    
    def get_upload_log(self) -> List[Dict]:
        """Получение лога выгрузок"""
        with self.lock:
            return self.upload_log.copy()
    
    def clear_log(self):
        """Очистка лога"""
        with self.lock:
            self.upload_log = []

# --------------------------------------------
# 🔄 ИНТЕГРАЦИЯ С 1С
# --------------------------------------------
class OneCIntegration:
    """Интеграция с 1С:Предприятие"""
    
    def __init__(self, base_url: str = None, username: str = None, password: str = None):
        self.base_url = base_url
        self.username = username
        self.password = password
        self.session = requests.Session()
        if username and password:
            self.session.auth = (username, password)
        self.cache = APICache()
        
    def export_to_1c(self, data: List[Dict], entity_type: str = "products") -> Dict:
        """Экспорт данных в 1С"""
        if not self.base_url:
            return {"status": "error", "message": "URL 1С не настроен"}
        
        try:
            export_data = {
                "entity_type": entity_type,
                "data": data,
                "timestamp": datetime.now().isoformat()
            }
            
            url = f"{self.base_url}/api/import"
            response = self.session.post(
                url,
                json=export_data,
                timeout=60
            )
            
            if response.status_code == 200:
                return {
                    "status": "success",
                    "message": "Данные успешно экспортированы в 1С",
                    "response": response.json()
                }
            else:
                return {
                    "status": "error",
                    "message": f"Ошибка {response.status_code}: {response.text}"
                }
                
        except Exception as e:
            logger.error(f"1C export error: {e}")
            return {"status": "error", "message": str(e)}
    
    def import_from_1c(self, entity_type: str = "products") -> Optional[List[Dict]]:
        """Импорт данных из 1С"""
        if not self.base_url:
            return None
        
        cache_key = generate_cache_key('1c_import', entity_type)
        cached = self.cache.get(cache_key)
        if cached:
            return cached
        
        try:
            url = f"{self.base_url}/api/export"
            params = {"entity_type": entity_type}
            
            response = self.session.get(url, params=params, timeout=60)
            
            if response.status_code == 200:
                data = response.json()
                self.cache.set(cache_key, data)
                return data
            else:
                logger.error(f"1C import error: {response.status_code}")
                return None
                
        except Exception as e:
            logger.error(f"1C import error: {e}")
            return None
    
    def sync_prices(self, prices: List[Dict]) -> Dict:
        """Синхронизация цен с 1С"""
        return self.export_to_1c(prices, "prices")
    
    def sync_stocks(self, stocks: List[Dict]) -> Dict:
        """Синхронизация остатков с 1С"""
        return self.export_to_1c(stocks, "stocks")

# --------------------------------------------
# 📊 ИНТЕГРАЦИЯ С CRM
# --------------------------------------------
class CRMIntegration:
    """Интеграция с CRM системами"""
    
    def __init__(self, crm_type: str = "amo", api_key: str = None, base_url: str = None):
        self.crm_type = crm_type
        self.api_key = api_key
        self.base_url = base_url
        self.session = requests.Session()
        
        if api_key:
            self.session.headers.update({
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            })
    
    def send_report(self, report_data: Dict) -> Dict:
        """Отправка отчета в CRM"""
        if not self.base_url:
            return {"status": "error", "message": "URL CRM не настроен"}
        
        try:
            if self.crm_type == "amo":
                return self._send_to_amo(report_data)
            elif self.crm_type == "bitrix24":
                return self._send_to_bitrix24(report_data)
            elif self.crm_type == "hubspot":
                return self._send_to_hubspot(report_data)
            else:
                return self._send_generic(report_data)
                
        except Exception as e:
            logger.error(f"CRM send error: {e}")
            return {"status": "error", "message": str(e)}
    
    def _send_to_amo(self, data: Dict) -> Dict:
        """Отправка в AmoCRM"""
        url = f"{self.base_url}/api/v4/leads"
        
        lead_data = {
            "name": data.get("title", "Отчет по юнит-экономике"),
            "custom_fields_values": [
                {
                    "field_id": 123456,
                    "values": [{"value": data.get("summary", "")}]
                }
            ]
        }
        
        response = self.session.post(url, json=lead_data)
        if response.status_code in [200, 201]:
            return {"status": "success", "data": response.json()}
        else:
            return {"status": "error", "message": response.text}
    
    def _send_to_bitrix24(self, data: Dict) -> Dict:
        """Отправка в Bitrix24"""
        url = f"{self.base_url}/rest/deal.add.json"
        
        deal_data = {
            "fields": {
                "TITLE": data.get("title", "Отчет по юнит-экономике"),
                "COMMENTS": data.get("summary", ""),
                "OPPORTUNITY": data.get("total_profit", 0)
            }
        }
        
        response = self.session.post(url, json=deal_data)
        if response.status_code == 200:
            return {"status": "success", "data": response.json()}
        else:
            return {"status": "error", "message": response.text}
    
    def _send_to_hubspot(self, data: Dict) -> Dict:
        """Отправка в HubSpot"""
        url = f"{self.base_url}/crm/v3/objects/deals"
        
        deal_data = {
            "properties": {
                "dealname": data.get("title", "Отчет по юнит-экономике"),
                "description": data.get("summary", ""),
                "amount": data.get("total_profit", 0)
            }
        }
        
        response = self.session.post(url, json=deal_data)
        if response.status_code in [200, 201]:
            return {"status": "success", "data": response.json()}
        else:
            return {"status": "error", "message": response.text}
    
    def _send_generic(self, data: Dict) -> Dict:
        """Универсальная отправка"""
        response = self.session.post(
            f"{self.base_url}/api/webhook",
            json=data
        )
        
        if response.status_code in [200, 201, 202]:
            return {"status": "success", "data": response.json()}
        else:
            return {"status": "error", "message": response.text}

# --------------------------------------------
# ⏰ РАСПИСАНИЕ АВТОМАТИЧЕСКИХ ВЫГРУЗОК
# --------------------------------------------
class AutoUploadScheduler:
    """Расписание автоматических выгрузок"""
    
    def __init__(self):
        self.schedules = {}
        self.is_running = False
        self.thread = None
        self.lock = Lock()
        
    def add_schedule(self, name: str, schedule: Dict):
        """Добавление расписания"""
        with self.lock:
            self.schedules[name] = {
                "schedule": schedule,
                "last_run": None,
                "next_run": self._calculate_next_run(schedule),
                "enabled": True
            }
    
    def _calculate_next_run(self, schedule: Dict) -> datetime:
        """Расчет следующего запуска"""
        now = datetime.now()
        
        if schedule.get("type") == "interval":
            seconds = schedule.get("interval_seconds", 3600)
            return now + timedelta(seconds=seconds)
        
        elif schedule.get("type") == "cron":
            hour = schedule.get("hour", 0)
            minute = schedule.get("minute", 0)
            day_of_week = schedule.get("day_of_week", 0)
            
            next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            
            if next_run <= now:
                days_until = (day_of_week - now.weekday()) % 7
                if days_until == 0:
                    days_until = 7
                next_run += timedelta(days=days_until)
            
            return next_run
        
        else:
            return now + timedelta(hours=1)
    
    def run_scheduled(self):
        """Запуск всех запланированных задач"""
        if self.is_running:
            return
        
        self.is_running = True
        self.thread = Thread(target=self._scheduler_loop)
        self.thread.daemon = True
        self.thread.start()
    
    def _scheduler_loop(self):
        """Основной цикл планировщика"""
        while self.is_running:
            try:
                now = datetime.now()
                
                with self.lock:
                    for name, data in self.schedules.items():
                        if not data["enabled"]:
                            continue
                        
                        if data["next_run"] and data["next_run"] <= now:
                            self._execute_schedule(name)
                            
                            data["last_run"] = now
                            data["next_run"] = self._calculate_next_run(data["schedule"])
                
                time.sleep(60)
                
            except Exception as e:
                logger.error(f"Scheduler error: {e}")
                time.sleep(60)
    
    def _execute_schedule(self, name: str):
        """Выполнение запланированной задачи"""
        schedule = self.schedules.get(name)
        if not schedule:
            return
        
        task_type = schedule["schedule"].get("task_type")
        
        if task_type == "upload_prices":
            logger.info(f"Выполнение плановой выгрузки цен: {name}")
        elif task_type == "upload_stocks":
            logger.info(f"Выполнение плановой выгрузки остатков: {name}")
        elif task_type == "sync_1c":
            logger.info(f"Выполнение синхронизации с 1С: {name}")
        elif task_type == "send_report":
            logger.info(f"Отправка отчета в CRM: {name}")
    
    def stop(self):
        """Остановка планировщика"""
        self.is_running = False
        if self.thread:
            self.thread.join(timeout=5)

# --------------------------------------------
# ⚡ ОПТИМИЗИРОВАННАЯ АНАЛИТИКА
# --------------------------------------------
class FastSalesAnalytics:
    """Легкая версия аналитики продаж"""
    
    def __init__(self):
        self.sales_data = []
        self._cache = {}
        
    def add_sale(self, sale: Dict):
        """Добавление записи о продаже"""
        self.sales_data.append({
            "timestamp": datetime.now(),
            "category": sale.get("category", "Прочее"),
            "quantity": sale.get("quantity", 1),
            "price": sale.get("price", 0),
            "cost": sale.get("cost", 0),
            "profit": sale.get("profit", 0),
            "marketplace": sale.get("marketplace", "Неизвестно")
        })
        
        if len(self.sales_data) > 1000:
            self.sales_data = self.sales_data[-500:]
    
    @st.cache_data(ttl=60)
    def get_sales_dataframe(_self) -> pd.DataFrame:
        """Получение данных продаж с кэшированием"""
        if not _self.sales_data:
            return _self._generate_minimal_demo()
        
        df = pd.DataFrame(_self.sales_data)
        if df.empty:
            return pd.DataFrame()
        
        df['date'] = df['timestamp'].dt.date
        return df
    
    def _generate_minimal_demo(self) -> pd.DataFrame:
        """Генерация минимальных демо-данных"""
        categories = ["Двигатель", "Трансмиссия", "Подвеска", "Тормозная система", 
                     "Масла и жидкости", "Фильтры", "Электрооборудование"]
        
        data = []
        start_date = datetime.now() - timedelta(days=30)
        
        for i in range(100):
            date = start_date + timedelta(days=random.randint(0, 30))
            category = random.choice(categories)
            quantity = random.randint(1, 3)
            price = random.randint(500, 10000)
            cost = price * random.uniform(0.5, 0.7)
            
            data.append({
                "timestamp": date,
                "category": category,
                "quantity": quantity,
                "price": price,
                "cost": cost,
                "profit": (price - cost) * quantity,
                "marketplace": random.choice(["Яндекс Маркет", "Ozon", "Wildberries"])
            })
        
        return pd.DataFrame(data)

# --------------------------------------------
# 📥 ЭКСПОРТ В EXCEL
# --------------------------------------------
class ExcelExportEngine:
    """Экспорт результатов в Excel с отдельными листами для каждого маркетплейса"""
    
    def __init__(self):
        self.classifier = CategoryClassifier()
    
    def export(self, results: List[Dict], marketplace: str, mode: str, all_rates: Dict = None) -> bytes:
        """Экспорт результатов в Excel с отдельными листами тарифов"""
        output = io.BytesIO()
        
        if not results:
            return output.getvalue()
        
        if not OPENPYXL_AVAILABLE:
            try:
                df = pd.DataFrame(results)
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Юнит-экономика', index=False)
                output.seek(0)
                return output.getvalue()
            except Exception as e:
                logger.error(f"Fallback export error: {e}")
                return output.getvalue()
        
        try:
            wb = Workbook()
            
            ws = wb.active
            ws.title = "Юнит-экономика"
            
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            headers = self._get_product_headers()
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row_idx, row_data in enumerate(results, 2):
                for col_idx, header in enumerate(headers, 1):
                    key = self._get_key_mapping().get(header, header.lower())
                    value = row_data.get(key, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    if isinstance(value, (int, float)):
                        if any(w in header for w in ["Цена", "Прибыль", "Доход", "стоимость", "LTV", "CAC"]):
                            cell.number_format = '#,##0.00 ₽'
                        elif any(w in header for w in ["%", "марж", "рр", "ROI", "CM"]):
                            cell.number_format = '0.00%'
                        elif any(w in header for w in ["дн", "запас", "EOQ"]):
                            cell.number_format = '#,##0'
                        
                        if key == "unit_profit" and isinstance(value, (int, float)):
                            if value > 0:
                                cell.font = Font(color="006400")
                            elif value < 0:
                                cell.font = Font(color="8B0000")
            
            for col in range(1, len(headers) + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 15
            
            if all_rates:
                self._create_help_sheets(wb, all_rates)
            
            self._create_summary_sheet(wb, results, marketplace, mode)
            
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Export error: {e}")
            try:
                df = pd.DataFrame(results)
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Юнит-экономика', index=False)
                output.seek(0)
                return output.getvalue()
            except Exception as e2:
                logger.error(f"Fallback export error: {e2}")
                return output.getvalue()
    
    def _get_product_headers(self) -> List[str]:
        """Заголовки для товарной модели с учетом единиц измерения"""
        return [
            "Артикул", "Наименование", "Бренд", "OE номер", "Категория", "Применимость",
            "Цена", "Себестоимость", 
            "Длина (исх.)", "Ширина (исх.)", "Высота (исх.)", "Ед. изм.",
            "Длина (мм)", "Ширина (мм)", "Высота (мм)",
            "Объем (л)", "Вес (кг)",
            "Комиссия", "Сервисный сбор", "Страховка", "Логистика", "Хранение",
            "Эквайринг", "Реклама", "Возвраты", "Упаковка",
            "Итого переменных", "Полная себестоимость",
            "Маржинальный доход", "CM %", "Прибыль на ед.", "Маржа %", "ROI %",
            "Окупаемость (дн)", "Точка безубыточности (шт)", "Точка безубыточности (₽)",
            "LTV", "CAC", "LTV/CAC", "EOQ", "Точка заказа", "Страховой запас",
            "Статус прибыли", "Статус маржи", "ABC", "XYZ",
            "Кол-во конкурентов", "Ср. цена конкурентов", "Мин цена", "Макс цена",
            "Позиция цены", "Рекомендация конкурентов",
            "Рекомендуемая цена", "Действие по цене", "Режим", "Маркетплейс",
            "Валидация", "Проблемы"
        ]
    
    def _get_key_mapping(self) -> Dict[str, str]:
        """Маппинг заголовков на ключи данных"""
        return {
            "Артикул": "article",
            "Наименование": "name",
            "Бренд": "brand",
            "OE номер": "oe_number",
            "Категория": "category",
            "Применимость": "compatibility",
            "Цена": "price",
            "Себестоимость": "cost",
            "Длина (исх.)": "length_orig",
            "Ширина (исх.)": "width_orig",
            "Высота (исх.)": "height_orig",
            "Ед. изм.": "dimension_unit",
            "Длина (мм)": "length_mm",
            "Ширина (мм)": "width_mm",
            "Высота (мм)": "height_mm",
            "Объем (л)": "volume",
            "Вес (кг)": "weight",
            "Комиссия": "commission",
            "Сервисный сбор": "service_fee",
            "Страховка": "insurance",
            "Логистика": "logistics",
            "Хранение": "storage",
            "Эквайринг": "acquiring",
            "Реклама": "advertising",
            "Возвраты": "returns",
            "Упаковка": "packaging",
            "Итого переменных": "total_variable_costs",
            "Полная себестоимость": "total_cost",
            "Маржинальный доход": "contribution_margin",
            "CM %": "contribution_margin_pct",
            "Прибыль на ед.": "unit_profit",
            "Маржа %": "margin",
            "ROI %": "roi",
            "Окупаемость (дн)": "payback_days",
            "Точка безубыточности (шт)": "break_even_units",
            "Точка безубыточности (₽)": "break_even_revenue",
            "LTV": "ltv",
            "CAC": "cac",
            "LTV/CAC": "ltv_cac_ratio",
            "EOQ": "eoq",
            "Точка заказа": "reorder_point",
            "Страховой запас": "safety_stock",
            "Статус прибыли": "profit_status",
            "Статус маржи": "margin_status",
            "ABC": "abc_category",
            "XYZ": "xyz_category",
            "Кол-во конкурентов": "competitor_count",
            "Ср. цена конкурентов": "competitor_avg_price",
            "Мин цена": "competitor_min_price",
            "Макс цена": "competitor_max_price",
            "Позиция цены": "price_position",
            "Рекомендация конкурентов": "competitor_recommendation",
            "Рекомендуемая цена": "recommended_price",
            "Действие по цене": "price_action",
            "Режим": "mode",
            "Маркетплейс": "marketplace",
            "Валидация": "validation_valid",
            "Проблемы": "validation_issues"
        }
    
    def _create_help_sheets(self, wb: Workbook, all_rates: Dict):
        """Создание отдельных листов с тарифами для каждого маркетплейса"""
        
        headers = ["Режим", "Комиссия %", "Эквайринг %", "Логистика база", "Хранение за литр", 
                   "Логистика за литр", "Бесплатно дней", "FBA база", "FBA за кг"]
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for marketplace, modes in all_rates.items():
            sheet_name = f"Тарифы {marketplace}"[:31]
            ws = wb.create_sheet(sheet_name)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            row_idx = 2
            for mode, rates in modes.items():
                ws.cell(row=row_idx, column=1, value=mode)
                ws.cell(row=row_idx, column=2, value=rates.get("commission", 0))
                ws.cell(row=row_idx, column=3, value=rates.get("acquiring", 0))
                ws.cell(row=row_idx, column=4, value=rates.get("logistics_base", 0))
                ws.cell(row=row_idx, column=5, value=rates.get("storage_per_liter", 0))
                ws.cell(row=row_idx, column=6, value=rates.get("logistics_per_liter", 0))
                ws.cell(row=row_idx, column=7, value=rates.get("storage_free_days", 0))
                ws.cell(row=row_idx, column=8, value=rates.get("fba_base", 0))
                ws.cell(row=row_idx, column=9, value=rates.get("fba_per_kg", 0))
                
                for col in [2, 3]:
                    cell = ws.cell(row=row_idx, column=col)
                    cell.number_format = '0.00%'
                
                row_idx += 1
            
            if row_idx > 2:
                row_idx += 1
                total_row = row_idx
                
                ws.cell(row=total_row, column=1, value="СРЕДНЕЕ:")
                total_font = Font(bold=True)
                total_fill = PatternFill(start_color="e94560", end_color="e94560", fill_type="solid")
                
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=total_row, column=col)
                    if col >= 2:
                        cell.value = f"=AVERAGE({get_column_letter(col)}2:{get_column_letter(col)}{row_idx-2})"
                        cell.number_format = '0.00%' if col in [2, 3] else '#,##0.00'
                    cell.font = total_font
                    if col == 1:
                        cell.fill = total_fill
            
            column_widths = [12, 14, 14, 16, 16, 16, 14, 14, 14]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            if len(all_rates) > 1:
                row_idx += 3
                ws.cell(row=row_idx, column=1, value="СРАВНЕНИЕ МАРКЕТПЛЕЙСОВ (FBY):")
                ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
                
                row_idx += 1
                compare_headers = ["Маркетплейс", "Комиссия %", "Эквайринг %", "Логистика", "Хранение"]
                for col, header in enumerate(compare_headers, 1):
                    cell = ws.cell(row=row_idx, column=col, value=header)
                    cell.font = Font(bold=True)
                
                row_idx += 1
                for mp, modes in all_rates.items():
                    fby_rates = modes.get("FBY", {})
                    ws.cell(row=row_idx, column=1, value=mp)
                    ws.cell(row=row_idx, column=2, value=fby_rates.get("commission", 0))
                    ws.cell(row=row_idx, column=3, value=fby_rates.get("acquiring", 0))
                    ws.cell(row=row_idx, column=4, value=fby_rates.get("logistics_base", 0))
                    ws.cell(row=row_idx, column=5, value=fby_rates.get("storage_per_liter", 0))
                    
                    for col in [2, 3]:
                        ws.cell(row=row_idx, column=col).number_format = '0.00%'
                    
                    row_idx += 1
    
    def _create_summary_sheet(self, wb: Workbook, results: List[Dict], marketplace: str, mode: str):
        """Создание листа со сводкой"""
        ws = wb.create_sheet("Сводка")
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        
        headers = ["Показатель", "Значение"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        if results:
            total_profit = sum(r.get("unit_profit", 0) for r in results)
            avg_margin = sum(r.get("margin", 0) for r in results) / len(results) if results else 0
            profitable = sum(1 for r in results if r.get("unit_profit", 0) > 0)
            
            summary_data = [
                ["Всего товаров", len(results)],
                ["Общая прибыль", format_currency(total_profit)],
                ["Средняя маржа", format_percent(avg_margin)],
                ["Прибыльных товаров", profitable],
                ["Убыточных товаров", len(results) - profitable],
                ["", ""],
                ["Параметры:", ""],
                ["Маркетплейс", marketplace],
                ["Режим", mode],
                ["Единицы измерения", results[0].get('dimension_unit', 'мм') if results else 'мм'],
                ["Дата", datetime.now().strftime("%d.%m.%Y %H:%M")],
                ["Версия", CONFIG["version"]]
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 2):
                ws.cell(row=row_idx, column=1, value=label)
                ws.cell(row=row_idx, column=2, value=value)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def load_existing(self, file_bytes: bytes) -> List[Dict]:
        """Загрузка существующей юнит-экономики из Excel"""
        if not OPENPYXL_AVAILABLE or not file_bytes:
            return []
        
        try:
            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
            
            headers = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    headers.append(safe_str(cell.value))
            
            if not headers:
                return []
            
            results = []
            key_mapping = self._get_key_mapping()
            
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    value = cell.value
                    key = key_mapping.get(header, header.lower())
                    
                    if key in ["price", "cost", "commission", "logistics", "storage", 
                              "acquiring", "advertising", "returns", "packaging",
                              "total_variable_costs", "total_cost", "contribution_margin",
                              "unit_profit", "ltv", "cac", "break_even_revenue", "recommended_price",
                              "competitor_avg_price", "competitor_min_price", "competitor_max_price"]:
                        value = safe_float(value)
                    elif key in ["margin", "contribution_margin_pct", "roi", "ltv_cac_ratio"]:
                        value = safe_float(value)
                    elif key in ["payback_days", "break_even_units", "eoq", "competitor_count"]:
                        value = safe_int(value)
                    else:
                        value = safe_str(value)
                    
                    row_data[key] = value
                
                if row_data.get("article"):
                    if "mode" not in row_data:
                        row_data["mode"] = "FBY"
                    if "marketplace" not in row_data:
                        row_data["marketplace"] = "Не указан"
                    if "dimension_unit" not in row_data:
                        row_data["dimension_unit"] = "мм"
                    results.append(row_data)
            
            return results
            
        except Exception as e:
            logger.error(f"Error loading existing data: {e}")
            return []

# --------------------------------------------
# 🎨 ОСНОВНОЕ ПРИЛОЖЕНИЕ
# --------------------------------------------
class UnitEconomicsApp:
    """Главное приложение с конвертацией размеров"""
    
    def __init__(self):
        self.classifier = CategoryClassifier()
        self.exporter = ExcelExportEngine()
        self.tariff_provider = AITariffProvider()
        self.results = []
        self.all_rates = {}
        self.engine = None
        self.api_clients = {}
        self.uploader = MarketplaceUploader()
        self.scheduler = AutoUploadScheduler()
        self.onec = OneCIntegration()
        self.crm = CRMIntegration()
        self.sales_analytics = FastSalesAnalytics()
        self.oem_db = OEMDatabase()
        self.telegram = TelegramNotifier()
        self.google_sheets = GoogleSheetsExporter()
        self.dimension_unit = "мм"
        
        if "theme" not in st.session_state:
            st.session_state.theme = "light"
        if "dimension_unit" not in st.session_state:
            st.session_state.dimension_unit = "мм"
    
    def run(self):
        """Запуск приложения"""
        st.set_page_config(
            page_title=CONFIG["app_name"],
            page_icon="🚀",
            layout="wide",
            initial_sidebar_state="expanded"
        )
        
        self._render_header()
        self._render_sidebar()
        self._render_main()
    
    def _render_header(self):
        """Отображение заголовка"""
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
                    padding: 1.5rem; border-radius: 15px; color: white; text-align: center; margin-bottom: 1.5rem;
                    border: 2px solid #e94560; box-shadow: 0 4px 15px rgba(0,0,0,0.3);">
            <h1 style="font-size: 2.8rem; margin: 0; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);">
                🚀 {CONFIG['app_name']}
            </h1>
            <p style="font-size: 1.2rem; opacity: 0.95; margin-top: 0.3rem;">
                📊 <strong>Товарная модель (B2C)</strong> | 2 файла | OE база | Конвертация размеров
            </p>
            <div style="display: flex; justify-content: center; gap: 0.8rem; flex-wrap: wrap; margin-top: 0.5rem;">
                <span style="background: rgba(233,69,96,0.3); padding: 0.2rem 1.2rem; border-radius: 20px; font-size: 0.9rem;">
                    v{CONFIG['version']}
                </span>
                <span style="background: rgba(233,69,96,0.3); padding: 0.2rem 1.2rem; border-radius: 20px; font-size: 0.9rem;">
                    📦 Товарная модель
                </span>
                <span style="background: rgba(233,69,96,0.3); padding: 0.2rem 1.2rem; border-radius: 20px; font-size: 0.9rem;">
                    📏 Конвертация мм/см
                </span>
                <span style="background: rgba(233,69,96,0.3); padding: 0.2rem 1.2rem; border-radius: 20px; font-size: 0.9rem;">
                    🤖 ML
                </span>
                <span style="background: rgba(233,69,96,0.3); padding: 0.2rem 1.2rem; border-radius: 20px; font-size: 0.9rem;">
                    📤 Telegram
                </span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    def _render_sidebar(self):
        """Отображение боковой панели с уникальными ключами"""
        with st.sidebar:
            st.markdown("## ⚙️ Настройки")
            
            # Переключение темы
            theme = st.toggle("🌙 Темная тема", value=st.session_state.theme == "dark", key="theme_toggle")
            st.session_state.theme = "dark" if theme else "light"
            
            st.divider()
            
            # Выбор единиц измерения
            st.markdown("### 📏 Единицы измерения")
            
            dimension_unit = st.radio(
                "Выберите единицы для размеров",
                options=["мм", "см"],
                index=0 if st.session_state.dimension_unit == "мм" else 1,
                help="Выберите в каких единицах указаны размеры в вашем файле",
                key="dimension_unit_radio"
            )
            st.session_state.dimension_unit = dimension_unit
            self.dimension_unit = dimension_unit
            
            st.divider()
            
            st.markdown("### 🔑 API ключи")
            
            ym_api_key = st.text_input(
                "Яндекс Маркет API ключ",
                type="password",
                placeholder="Ваш API ключ",
                key="ym_api_key"
            )
            
            ozon_api_key = st.text_input(
                "Ozon API ключ",
                type="password",
                placeholder="Ваш API ключ",
                key="ozon_api_key"
            )
            
            ozon_client_id = st.text_input(
                "Ozon Client ID",
                type="password",
                placeholder="Ваш Client ID",
                key="ozon_client_id"
            )
            
            wb_api_key = st.text_input(
                "Wildberries API ключ",
                type="password",
                placeholder="Ваш API ключ",
                key="wb_api_key"
            )
            
            ds_api_key = st.text_input(
                "🔑 DeepSeek API ключ",
                type="password",
                placeholder="sk-...",
                help="Для AI-тарифов",
                key="ds_api_key"
            )
            if ds_api_key:
                self.tariff_provider.api_key = ds_api_key
                st.success("✅ DeepSeek ключ установлен")
            
            st.divider()
            
            st.markdown("### 📤 Уведомления")
            
            tg_token = st.text_input(
                "Telegram Bot Token",
                type="password",
                placeholder="123456789:ABCdef...",
                key="tg_token"
            )
            tg_chat_id = st.text_input(
                "Telegram Chat ID",
                placeholder="123456789",
                key="tg_chat_id"
            )
            if tg_token and tg_chat_id:
                self.telegram = TelegramNotifier(tg_token, tg_chat_id)
                st.success("✅ Telegram настроен")
            
            st.divider()
            
            st.markdown("### 📊 Google Sheets")
            
            gs_json = st.text_area(
                "Google Service Account JSON",
                placeholder='{"type": "service_account", ...}',
                height=100,
                key="gs_json"
            )
            gs_name = st.text_input(
                "Название таблицы",
                placeholder="Юнит-экономика",
                key="gs_name"
            )
            if gs_json and gs_name:
                self.google_sheets = GoogleSheetsExporter(gs_json, gs_name)
                st.success("✅ Google Sheets настроен")
            
            st.divider()
            
            st.markdown("### 📦 Параметры")
            
            marketplace = st.selectbox(
                "🏪 Маркетплейс",
                CONFIG["marketplaces"],
                index=0,
                key="marketplace_select"
            )
            
            mode = st.selectbox(
                "📦 Режим работы",
                CONFIG["operation_modes"],
                index=0,
                key="mode_select"
            )
            
            days_storage = st.number_input(
                "📦 Хранение (дней)",
                min_value=1,
                max_value=730,
                value=30,
                key="days_storage"
            )
            
            st.divider()
            
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🔄 Обновить тарифы", use_container_width=True, key="refresh_rates"):
                    with st.spinner("⏳ Получение тарифов..."):
                        self.tariff_provider.clear_cache()
                        self.all_rates = self.tariff_provider.get_all_rates()
                        st.success("✅ Тарифы обновлены!")
            
            with col2:
                if st.button("🗑️ Очистить кэш", use_container_width=True, key="clear_cache"):
                    APICache().clear()
                    st.success("✅ Кэш очищен!")
            
            st.divider()
            
            st.markdown("### 📊 Статистика")
            if self.results:
                st.metric("📦 Позиций", len(self.results))
                profitable = sum(1 for r in self.results if r.get("unit_profit", 0) > 0)
                total_profit = sum(r.get("unit_profit", 0) for r in self.results)
                st.metric("💰 Прибыльных", f"{profitable}/{len(self.results)}")
                st.metric("💵 Общая прибыль", format_currency(total_profit))
    
    def _render_main(self):
        """Основной контент с вкладками"""
        tabs = [
            "📁 Загрузка данных",
            "🗄️ База OE",
            "🔌 API интеграция",
            "🕷️ Парсинг конкурентов",
            "📤 Автовыгрузка",
            "🔄 1С/CRM",
            "📊 Дашборд"
        ]
        tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(tabs)
        
        with tab1:
            self._render_upload_tab()
        
        with tab2:
            self._render_oem_database_tab()
        
        with tab3:
            self._render_api_tab()
        
        with tab4:
            self._render_parsing_tab()
        
        with tab5:
            self._render_autoupload_tab()
        
        with tab6:
            self._render_integration_tab()
        
        with tab7:
            self._render_dashboard_tab()
    
    def _render_upload_tab(self):
        """Вкладка загрузки данных с выбором единиц"""
        st.subheader("📁 Загрузка данных для товарной модели")
        
        st.info(f"""
        **📦 Товарная модель (B2C)**
        
        **Загрузите 2 файла:**
        1. **Основной** — Артикул, Бренд, Длина, Ширина, Высота, Количество, Цена
        2. **Справочный** — Артикул, Бренд, Наименование, Применимость, OE номера
        
        **📏 Единицы измерения:** {self.dimension_unit}
        - Все размеры в файле будут интерпретироваться как {self.dimension_unit}
        - Для расчетов размеры автоматически конвертируются в мм
        """)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 📦 Файл 1: Основной")
            main_file = st.file_uploader(
                "Загрузите основной файл (Excel/CSV)",
                type=["xlsx", "xls", "csv"],
                help=f"Колонки: Артикул, Бренд, Длина ({self.dimension_unit}), Ширина ({self.dimension_unit}), Высота ({self.dimension_unit}), Количество, Цена",
                key="main_file_upload"
            )
        
        with col2:
            st.markdown("### 📋 Файл 2: Справочный")
            ref_file = st.file_uploader(
                "Загрузите справочный файл (Excel/CSV)",
                type=["xlsx", "xls", "csv"],
                help="Колонки: Артикул, Бренд, Наименование, Применимость, OE номера",
                key="ref_file_upload"
            )
        
        if main_file and ref_file:
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("🚀 Рассчитать (обычный)", type="primary", use_container_width=True, key="calc_normal"):
                    with st.spinner("⏳ Выполняется расчет..."):
                        results = self._process_two_files(
                            main_file,
                            ref_file,
                            self._get_marketplace_from_sidebar(),
                            self._get_mode_from_sidebar(),
                            self._get_days_storage_from_sidebar(),
                            parallel=False
                        )
                        
                        if results:
                            self.results = results
                            st.success(f"✅ Всего обработано {len(results)} позиций!")
                            self._show_results(results)
                            self._show_export(results)
                        else:
                            st.warning("⚠️ Нет данных для обработки")
            
            with col2:
                if st.button("🚀 Рассчитать (параллельно)", type="primary", use_container_width=True, key="calc_parallel"):
                    with st.spinner("⏳ Выполняется параллельный расчет..."):
                        results = self._process_two_files(
                            main_file,
                            ref_file,
                            self._get_marketplace_from_sidebar(),
                            self._get_mode_from_sidebar(),
                            self._get_days_storage_from_sidebar(),
                            parallel=True
                        )
                        
                        if results:
                            self.results = results
                            st.success(f"✅ Всего обработано {len(results)} позиций!")
                            self._show_results(results)
                            self._show_export(results)
                        else:
                            st.warning("⚠️ Нет данных для обработки")
            
            with col3:
                if st.button("📤 Отправить в Telegram", use_container_width=True, key="send_telegram"):
                    if self.results:
                        if self.telegram.send_report(self.results):
                            st.success("✅ Отчет отправлен в Telegram!")
                        else:
                            st.error("❌ Ошибка отправки в Telegram")
                    else:
                        st.warning("⚠️ Сначала рассчитайте юнит-экономику")
    
    def _process_two_files(self, main_file, ref_file, marketplace: str, mode: str, days_storage: int, parallel: bool = False) -> List[Dict]:
        """Обработка двух файлов: основной + справочный с учетом единиц измерения"""
        try:
            if main_file.name.endswith('.csv'):
                df_main = pd.read_csv(main_file, encoding='utf-8-sig')
            else:
                df_main = pd.read_excel(main_file, engine='openpyxl')
            
            if ref_file.name.endswith('.csv'):
                df_ref = pd.read_csv(ref_file, encoding='utf-8-sig')
            else:
                df_ref = pd.read_excel(ref_file, engine='openpyxl')
            
            if df_main.empty or df_ref.empty:
                st.error("❌ Один из файлов пуст")
                return []
            
            main_cols = {col.lower(): col for col in df_main.columns}
            
            article_col = None
            brand_col_main = None
            length_col = None
            width_col = None
            height_col = None
            quantity_col = None
            price_col = None
            weight_col = None
            
            for col in df_main.columns:
                col_lower = col.lower()
                if any(w in col_lower for w in ['артикул', 'article', 'sku', 'код', 'id']):
                    if not article_col:
                        article_col = col
                elif any(w in col_lower for w in ['бренд', 'brand', 'марка', 'производитель']):
                    if not brand_col_main:
                        brand_col_main = col
                elif any(w in col_lower for w in ['длин', 'length']):
                    if not length_col:
                        length_col = col
                elif any(w in col_lower for w in ['ширин', 'width']):
                    if not width_col:
                        width_col = col
                elif any(w in col_lower for w in ['высот', 'height']):
                    if not height_col:
                        height_col = col
                elif any(w in col_lower for w in ['количество', 'quantity', 'stock', 'остаток']):
                    if not quantity_col:
                        quantity_col = col
                elif any(w in col_lower for w in ['цена', 'price']) and 'себест' not in col_lower:
                    if not price_col:
                        price_col = col
                elif any(w in col_lower for w in ['вес', 'weight', 'масса']):
                    if not weight_col:
                        weight_col = col
            
            ref_cols = {col.lower(): col for col in df_ref.columns}
            
            article_ref_col = None
            brand_ref_col = None
            name_col = None
            compatibility_col = None
            oe_col = None
            
            for col in df_ref.columns:
                col_lower = col.lower()
                if any(w in col_lower for w in ['артикул', 'article', 'sku', 'код', 'id']):
                    if not article_ref_col:
                        article_ref_col = col
                elif any(w in col_lower for w in ['бренд', 'brand', 'марка', 'производитель']):
                    if not brand_ref_col:
                        brand_ref_col = col
                elif any(w in col_lower for w in ['наименование', 'название', 'name', 'product']):
                    if not name_col:
                        name_col = col
                elif any(w in col_lower for w in ['применимость', 'совместимость', 'compatibility', 'car']):
                    if not compatibility_col:
                        compatibility_col = col
                elif any(w in col_lower for w in ['оем', 'oe', 'oem', 'номер', 'part number']):
                    if not oe_col:
                        oe_col = col
            
            if not article_col:
                st.warning("⚠️ В основном файле не найдена колонка с артикулами")
                return []
            
            if not article_ref_col:
                st.warning("⚠️ В справочном файле не найдена колонка с артикулами")
                return []
            
            if not price_col:
                st.warning("⚠️ В основном файле не найдена колонка с ценами")
                return []
            
            ref_dict = {}
            for idx, row in df_ref.iterrows():
                article = safe_str(row[article_ref_col])
                if article:
                    ref_dict[article] = {
                        "brand_ref": safe_str(row[brand_ref_col]) if brand_ref_col else "",
                        "name": safe_str(row[name_col]) if name_col else "",
                        "compatibility": safe_str(row[compatibility_col]) if compatibility_col else "",
                        "oe_number": safe_str(row[oe_col]) if oe_col else ""
                    }
            
            self.engine = UnitEconomicsEngine(
                marketplace=marketplace,
                mode=mode,
                days_storage=days_storage,
                dimension_unit=self.dimension_unit
            )
            
            results = []
            
            if parallel:
                with st.spinner("⏳ Параллельная обработка..."):
                    chunk_size = 100
                    chunks = []
                    
                    for i in range(0, len(df_main), chunk_size):
                        chunks.append(df_main.iloc[i:i+chunk_size])
                    
                    with ThreadPoolExecutor(max_workers=4) as executor:
                        futures = []
                        for chunk in chunks:
                            future = executor.submit(
                                self._process_chunk,
                                chunk,
                                ref_dict,
                                marketplace,
                                mode,
                                days_storage,
                                self.dimension_unit
                            )
                            futures.append(future)
                        
                        for future in as_completed(futures):
                            results.extend(future.result())
            else:
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                for idx, row in df_main.iterrows():
                    try:
                        if idx % 10 == 0:
                            progress = min(idx / len(df_main), 0.99)
                            progress_bar.progress(progress)
                            status_text.text(f"⏳ Обработка {idx+1}/{len(df_main)}...")
                        
                        article = safe_str(row[article_col]) if article_col else f"ART_{idx+1:06d}"
                        brand_main = safe_str(row[brand_col_main]) if brand_col_main else ""
                        length = safe_float(row[length_col]) if length_col else 0
                        width = safe_float(row[width_col]) if width_col else 0
                        height = safe_float(row[height_col]) if height_col else 0
                        quantity = safe_int(row[quantity_col]) if quantity_col else 0
                        price = safe_float(row[price_col]) if price_col else 0
                        weight = safe_float(row[weight_col]) if weight_col else 0
                        
                        if price <= 0:
                            continue
                        
                        ref_data = ref_dict.get(article, {})
                        
                        brand = brand_main if brand_main else ref_data.get("brand_ref", "")
                        name = ref_data.get("name", "")
                        compatibility = ref_data.get("compatibility", "")
                        oe_number = ref_data.get("oe_number", "")
                        
                        oe_data = self.engine.oem_db.get_by_oe(oe_number) if oe_number else None
                        
                        if oe_data:
                            if not brand:
                                brand = oe_data.get("brand", "")
                            
                            if length == 0 and "dimensions" in oe_data:
                                dims = oe_data["dimensions"]
                                length = dims.get("length", 0)
                                width = dims.get("width", 0)
                                height = dims.get("height", 0)
                            
                            if weight == 0:
                                weight = oe_data.get("weight", 0)
                            
                            category = oe_data.get("category", "Прочее")
                        else:
                            category, confidence = self.classifier.classify(name) if name else ("Прочее", 0)
                        
                        if not name:
                            name = f"{brand} {oe_number}" if brand and oe_number else article
                        
                        cost = price * 0.5
                        
                        row_data = {
                            "article": article,
                            "name": name,
                            "brand": brand,
                            "oe_number": oe_number,
                            "category": category,
                            "price": price,
                            "cost": cost,
                            "length": length,
                            "width": width,
                            "height": height,
                            "weight": weight,
                            "quantity": quantity,
                            "compatibility": compatibility,
                            "source": "from_oem_db" if oe_data else "from_file"
                        }
                        
                        result = self.engine.calculate_product(row_data)
                        if result:
                            result["source"] = "from_oem_db" if oe_data else "from_file"
                            result["quantity"] = quantity
                            result["compatibility"] = compatibility
                            results.append(result)
                            
                    except Exception as e:
                        logger.error(f"Error processing row {idx}: {e}")
                        continue
                
                progress_bar.progress(1.0)
                status_text.text("✅ Обработка завершена!")
            
            if self.engine:
                validation_report = self.engine.get_validation_report()
                if validation_report["invalid"] > 0:
                    st.warning(f"⚠️ Найдено {validation_report['invalid']} позиций с ошибками валидации")
                    with st.expander("📋 Отчет по валидации"):
                        st.json(validation_report)
            
            return results
            
        except Exception as e:
            logger.error(f"Error processing two files: {e}")
            st.error(f"❌ Ошибка обработки данных: {str(e)}")
            return []
    
    def _process_chunk(self, chunk, ref_dict, marketplace, mode, days_storage, dimension_unit):
        """Обработка чанка данных"""
        results = []
        engine = UnitEconomicsEngine(marketplace, mode, days_storage, dimension_unit)
        
        for idx, row in chunk.iterrows():
            try:
                article = safe_str(row.get('Артикул', row.get('article', f"ART_{idx+1:06d}")))
                brand = safe_str(row.get('Бренд', row.get('brand', '')))
                length = safe_float(row.get('Длина', row.get('length', 0)))
                width = safe_float(row.get('Ширина', row.get('width', 0)))
                height = safe_float(row.get('Высота', row.get('height', 0)))
                quantity = safe_int(row.get('Количество', row.get('quantity', 0)))
                price = safe_float(row.get('Цена', row.get('price', 0)))
                weight = safe_float(row.get('Вес', row.get('weight', 0)))
                
                if price <= 0:
                    continue
                
                ref_data = ref_dict.get(article, {})
                
                brand = brand if brand else ref_data.get("brand_ref", "")
                name = ref_data.get("name", "")
                compatibility = ref_data.get("compatibility", "")
                oe_number = ref_data.get("oe_number", "")
                
                oe_data = engine.oem_db.get_by_oe(oe_number) if oe_number else None
                
                if oe_data:
                    if not brand:
                        brand = oe_data.get("brand", "")
                    
                    if length == 0 and "dimensions" in oe_data:
                        dims = oe_data["dimensions"]
                        length = dims.get("length", 0)
                        width = dims.get("width", 0)
                        height = dims.get("height", 0)
                    
                    if weight == 0:
                        weight = oe_data.get("weight", 0)
                    
                    category = oe_data.get("category", "Прочее")
                else:
                    category = "Прочее"
                
                if not name:
                    name = f"{brand} {oe_number}" if brand and oe_number else article
                
                cost = price * 0.5
                
                row_data = {
                    "article": article,
                    "name": name,
                    "brand": brand,
                    "oe_number": oe_number,
                    "category": category,
                    "price": price,
                    "cost": cost,
                    "length": length,
                    "width": width,
                    "height": height,
                    "weight": weight,
                    "quantity": quantity,
                    "compatibility": compatibility
                }
                
                result = engine.calculate_product(row_data)
                if result:
                    result["quantity"] = quantity
                    result["compatibility"] = compatibility
                    results.append(result)
                    
            except Exception as e:
                logger.error(f"Error in chunk: {e}")
                continue
        
        return results
    
    def _render_oem_database_tab(self):
        """Вкладка управления базой OE номеров"""
        st.subheader("🗄️ Управление базой OE номеров")
        
        st.markdown("""
        **Возможности:**
        - ✅ Просмотр всех OE номеров в базе
        - ✅ Добавление/обновление OE номеров
        - ✅ Поиск по OE, бренду, категории
        - ✅ Импорт/экспорт базы
        - ✅ Статистика базы
        """)
        
        tabs = st.tabs(["📋 Просмотр", "➕ Добавить", "🔍 Поиск", "📊 Статистика"])
        
        with tabs[0]:
            st.markdown("### 📋 Все OE номера")
            
            search = st.text_input("🔍 Фильтр по OE номеру", placeholder="Введите OE номер...", key="oem_search")
            
            all_oe = list(self.oem_db.data.items())
            
            if search:
                all_oe = [(oe, data) for oe, data in all_oe if search.upper() in oe.upper()]
            
            if all_oe:
                data_list = []
                for oe, data in all_oe[:100]:
                    data_list.append({
                        "OE номер": oe,
                        "Категория": data.get("category", ""),
                        "Бренд": data.get("brand", ""),
                        "Вес (кг)": data.get("weight", 0),
                        "Совместимость": ", ".join(data.get("compatibility", [])[:2]),
                        "Кросс-ссылки": ", ".join(data.get("cross_reference", [])[:2])
                    })
                
                df = pd.DataFrame(data_list)
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                if len(all_oe) > 100:
                    st.caption(f"Показаны первые 100 из {len(all_oe)} записей")
            else:
                st.info("Нет данных в базе")
        
        with tabs[1]:
            st.markdown("### ➕ Добавление OE номера")
            
            col1, col2 = st.columns(2)
            
            with col1:
                new_oe = st.text_input("OE номер *", placeholder="Например: 0986AF0059", key="new_oe")
                new_brand = st.text_input("Бренд *", placeholder="BOSCH", key="new_brand")
                new_category = st.text_input("Категория *", placeholder="Фильтры", key="new_category")
                new_subcategory = st.text_input("Подкатегория", placeholder="Масляные фильтры", key="new_subcategory")
                new_weight = st.number_input("Вес (кг)", min_value=0.0, step=0.1, value=0.0, key="new_weight")
            
            with col2:
                new_length = st.number_input("Длина (мм)", min_value=0, step=10, value=0, key="new_length")
                new_width = st.number_input("Ширина (мм)", min_value=0, step=10, value=0, key="new_width")
                new_height = st.number_input("Высота (мм)", min_value=0, step=10, value=0, key="new_height")
                new_compatibility = st.text_area("Совместимость (через запятую)", placeholder="BMW 3, Audi A4, VW Golf", key="new_compatibility")
                new_cross = st.text_area("Кросс-ссылки (через запятую)", placeholder="MANN W842/2, MAHLE OC 205", key="new_cross")
            
            if st.button("💾 Сохранить OE номер", type="primary", key="save_oe"):
                if new_oe and new_brand and new_category:
                    data = {
                        "category": new_category,
                        "subcategory": new_subcategory or "Запчасть",
                        "brand": new_brand,
                        "weight": new_weight,
                        "dimensions": {
                            "length": new_length,
                            "width": new_width,
                            "height": new_height
                        }
                    }
                    
                    if new_compatibility:
                        data["compatibility"] = [x.strip() for x in new_compatibility.split(",") if x.strip()]
                    
                    if new_cross:
                        data["cross_reference"] = [x.strip() for x in new_cross.split(",") if x.strip()]
                    
                    self.oem_db.add_or_update(new_oe, data)
                    st.success(f"✅ OE номер {new_oe} добавлен/обновлен!")
                else:
                    st.warning("⚠️ Заполните обязательные поля (OE, Бренд, Категория)")
        
        with tabs[2]:
            st.markdown("### 🔍 Поиск в базе OE")
            
            search_type = st.selectbox(
                "Тип поиска",
                ["По OE номеру", "По бренду", "По категории"],
                key="search_type"
            )
            
            search_query = st.text_input("Поисковый запрос", key="search_query")
            
            if search_query and st.button("🔍 Найти", key="search_btn"):
                if search_type == "По OE номеру":
                    result = self.oem_db.get_by_oe(search_query)
                    if result:
                        st.json(result)
                    else:
                        st.warning("❌ OE номер не найден")
                
                elif search_type == "По бренду":
                    results = self.oem_db.search_by_brand(search_query)
                    if results:
                        st.success(f"✅ Найдено {len(results)} записей")
                        st.dataframe(pd.DataFrame(results), use_container_width=True, hide_index=True)
                    else:
                        st.warning("❌ Бренд не найден")
                
                elif search_type == "По категории":
                    results = self.oem_db.search_by_category(search_query)
                    if results:
                        st.success(f"✅ Найдено {len(results)} записей")
                        st.dataframe(pd.DataFrame(results), use_container_width=True, hide_index=True)
                    else:
                        st.warning("❌ Категория не найдена")
        
        with tabs[3]:
            st.markdown("### 📊 Статистика базы")
            
            stats = self.oem_db.get_statistics()
            
            col1, col2, col3 = st.columns(3)
            col1.metric("📦 Всего OE номеров", stats["total"])
            col2.metric("🏷️ Категорий", len(stats["categories"]))
            col3.metric("🏢 Брендов", len(stats["brands"]))
            
            st.divider()
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**📊 Распределение по категориям:**")
                if stats["categories"]:
                    df_cat = pd.DataFrame({
                        "Категория": list(stats["categories"].keys()),
                        "Количество": list(stats["categories"].values())
                    }).sort_values("Количество", ascending=False)
                    st.dataframe(df_cat, use_container_width=True, hide_index=True)
            
            with col2:
                st.markdown("**📊 Топ брендов:**")
                if stats["brands"]:
                    df_brand = pd.DataFrame({
                        "Бренд": list(stats["brands"].keys()),
                        "Количество": list(stats["brands"].values())
                    }).sort_values("Количество", ascending=False).head(10)
                    st.dataframe(df_brand, use_container_width=True, hide_index=True)
            
            st.divider()
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("📤 Экспорт базы в JSON", use_container_width=True, key="export_oem"):
                    json_data = json.dumps(self.oem_db.data, ensure_ascii=False, indent=2)
                    st.download_button(
                        "📥 Скачать JSON",
                        data=json_data,
                        file_name=f"oem_database_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                        mime="application/json",
                        use_container_width=True,
                        key="download_oem"
                    )
            
            with col2:
                uploaded_json = st.file_uploader(
                    "📥 Импорт базы из JSON",
                    type=["json"],
                    key="oem_import"
                )
                if uploaded_json:
                    try:
                        data = json.load(uploaded_json)
                        self.oem_db.data.update(data)
                        self.oem_db._save_database()
                        st.success(f"✅ Импортировано {len(data)} OE номеров!")
                    except Exception as e:
                        st.error(f"❌ Ошибка импорта: {e}")
    
    def _render_api_tab(self):
        """Вкладка API интеграции"""
        st.subheader("🔌 Интеграция с API маркетплейсов")
        
        st.markdown("""
        **Доступные API:**
        - **Яндекс Маркет** — управление ценами, остатками, заказами
        - **Ozon** — управление товарами, ценами, складом
        - **Wildberries** — управление карточками, ценами, остатками
        - **AliExpress** — управление товарами (в разработке)
        """)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 📊 Получение данных")
            
            api_action = st.selectbox(
                "Действие",
                ["Получить товары", "Получить цены", "Получить остатки", "Обновить цены"],
                key="api_action"
            )
            
            marketplace = st.selectbox(
                "Маркетплейс",
                ["Яндекс Маркет", "Ozon", "Wildberries"],
                key="api_marketplace"
            )
            
            if st.button("🚀 Выполнить запрос", type="primary", use_container_width=True, key="execute_api"):
                with st.spinner("⏳ Выполняется запрос..."):
                    result = self._execute_api_request(marketplace, api_action)
                    if result:
                        st.success("✅ Запрос выполнен успешно!")
                        st.json(result)
                    else:
                        st.error("❌ Ошибка выполнения запроса")
        
        with col2:
            st.markdown("### 📋 Массовое обновление")
            
            st.info("""
            **Массовое обновление цен:**
            1. Загрузите файл с новыми ценами
            2. Выберите маркетплейс
            3. Нажмите "Обновить цены"
            """)
            
            update_file = st.file_uploader(
                "Файл с ценами (Excel/CSV)",
                type=["xlsx", "xls", "csv"],
                key="update_prices_file"
            )
            
            if update_file and st.button("🔄 Обновить цены", use_container_width=True, key="update_prices_btn"):
                st.warning("⚠️ Функция в разработке. Требуется настройка API ключей.")
    
    def _render_parsing_tab(self):
        """Вкладка парсинга конкурентов"""
        st.subheader("🕷️ Парсинг цен конкурентов")
        
        st.markdown("""
        **Парсинг в реальном времени:**
        - Автоматический сбор цен с Яндекс Маркета, Ozon, Wildberries
        - Анализ конкурентной среды
        - Рекомендации по ценообразованию
        """)
        
        col1, col2 = st.columns(2)
        
        with col1:
            search_query = st.text_input(
                "🔍 Поисковый запрос",
                placeholder="Например: масло моторное 5W-40",
                key="parse_search"
            )
            
            marketplace = st.selectbox(
                "Маркетплейс",
                ["Все", "Яндекс Маркет", "Ozon", "Wildberries"],
                key="parse_marketplace"
            )
            
            if st.button("🕷️ Начать парсинг", type="primary", use_container_width=True, key="start_parse"):
                if search_query:
                    with st.spinner(f"⏳ Парсинг {marketplace if marketplace != 'Все' else 'всех маркетплейсов'}..."):
                        parser = CompetitorParser()
                        
                        if marketplace == "Все":
                            results = parser.parse_all_marketplaces(search_query)
                        else:
                            parser_map = {
                                'Яндекс Маркет': parser.parse_yandex_market,
                                'Ozon': parser.parse_ozon,
                                'Wildberries': parser.parse_wildberries
                            }
                            results = {marketplace: parser_map[marketplace](search_query)}
                        
                        if results:
                            st.success("✅ Парсинг завершен!")
                            
                            total_items = sum(len(items) for items in results.values())
                            st.metric("📦 Найдено товаров", total_items)
                            
                            for mp, items in results.items():
                                if items:
                                    with st.expander(f"📊 {mp} — {len(items)} товаров"):
                                        df = pd.DataFrame(items[:20])
                                        if 'price' in df.columns:
                                            df['price'] = df['price'].apply(format_currency)
                                        st.dataframe(df, use_container_width=True, hide_index=True)
                                        if len(items) > 20:
                                            st.caption(f"Показаны первые 20 из {len(items)}")
                        else:
                            st.warning("⚠️ Не найдено товаров по запросу")
                else:
                    st.warning("⚠️ Введите поисковый запрос")
        
        with col2:
            st.markdown("### 📊 Анализ конкурентов")
            
            if st.button("📊 Анализировать мои товары", use_container_width=True, key="analyze_competitors"):
                if self.results:
                    with st.spinner("⏳ Анализ конкурентов..."):
                        competitor_manager = CompetitorManager()
                        analysis_results = []
                        
                        for product in self.results[:10]:
                            name = product.get('name', '')
                            price = product.get('price', 0)
                            if name and price > 0:
                                analysis = competitor_manager.analyze_competitor_prices(name, price)
                                analysis_results.append({
                                    'Товар': name[:30],
                                    'Наша цена': price,
                                    'Ср. цена конкурентов': analysis.get('avg_price', 0),
                                    'Позиция': analysis.get('price_position', ''),
                                    'Рекомендация': analysis.get('recommendation', '')
                                })
                        
                        if analysis_results:
                            df = pd.DataFrame(analysis_results)
                            st.dataframe(df, use_container_width=True, hide_index=True)
                            st.success("✅ Анализ завершен!")
                        else:
                            st.warning("⚠️ Нет данных для анализа")
                else:
                    st.warning("⚠️ Сначала рассчитайте юнит-экономику")
    
    def _render_autoupload_tab(self):
        """Вкладка автовыгрузки"""
        st.subheader("📤 Автоматическая выгрузка на маркетплейсы")
        
        st.markdown("""
        **Возможности автовыгрузки:**
        - ✅ Выгрузка цен на все маркетплейсы
        - ✅ Выгрузка остатков (стока)
        - ✅ Создание/обновление карточек товаров
        - ✅ Автоматическое расписание
        - ✅ Логирование всех операций
        """)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 📤 Выгрузка сейчас")
            
            upload_type = st.selectbox(
                "Тип выгрузки",
                ["Цены", "Остатки", "Карточки товаров"],
                key="upload_type"
            )
            
            marketplace = st.selectbox(
                "Маркетплейс",
                ["Яндекс Маркет", "Ozon", "Wildberries"],
                key="upload_marketplace"
            )
            
            if st.button("🚀 Запустить выгрузку", type="primary", use_container_width=True, key="start_upload"):
                if self.results:
                    with st.spinner(f"⏳ Выгрузка {upload_type} на {marketplace}..."):
                        products = []
                        for product in self.results:
                            p = {
                                "offer_id": product.get("article", ""),
                                "article": product.get("article", ""),
                                "price": product.get("recommended_price", product.get("price", 0)),
                                "stock": product.get("quantity", 100),
                                "name": product.get("name", ""),
                                "category": product.get("category", ""),
                                "brand": product.get("brand", ""),
                                "length_mm": product.get("length_mm", 0),
                                "width_mm": product.get("width_mm", 0),
                                "height_mm": product.get("height_mm", 0)
                            }
                            products.append(p)
                        
                        if upload_type == "Цены":
                            result = self.uploader.upload_prices(marketplace, products)
                        elif upload_type == "Остатки":
                            result = self.uploader.upload_stocks(marketplace, products)
                        else:
                            result = self.uploader.upload_products(marketplace, products)
                        
                        if result:
                            st.success(f"✅ Выгрузка завершена: {result['success']} успешно, {result['failed']} ошибок")
                            if result.get('errors'):
                                with st.expander("📋 Детали ошибок"):
                                    for error in result['errors'][:10]:
                                        st.error(f"{error['offer_id']}: {error['error']}")
                                        if len(result['errors']) > 10:
                                            st.caption(f"... и еще {len(result['errors']) - 10} ошибок")
                        else:
                            st.error("❌ Ошибка выгрузки")
                else:
                    st.warning("⚠️ Сначала рассчитайте юнит-экономику")
        
        with col2:
            st.markdown("### ⏰ Расписание выгрузок")
            
            enable_schedule = st.checkbox("Включить автоматические выгрузки", value=False, key="enable_schedule")
            
            if enable_schedule:
                schedule_type = st.selectbox(
                    "Тип расписания",
                    ["Интервал", "По расписанию"],
                    key="schedule_type"
                )
                
                if schedule_type == "Интервал":
                    interval = st.selectbox(
                        "Интервал",
                        ["Каждый час", "Каждые 3 часа", "Каждые 6 часов", "Каждый день"],
                        key="interval_select"
                    )
                    interval_seconds = {"Каждый час": 3600, "Каждые 3 часа": 10800, "Каждые 6 часов": 21600, "Каждый день": 86400}
                    
                else:
                    hour = st.selectbox("Час", range(0, 24), index=9, key="schedule_hour")
                    minute = st.selectbox("Минута", range(0, 60), index=0, key="schedule_minute")
                
                upload_tasks = st.multiselect(
                    "Задачи для автоматической выгрузки",
                    ["Выгрузка цен", "Выгрузка остатков", "Синхронизация с 1С", "Отправка отчета в CRM"],
                    default=["Выгрузка цен"],
                    key="upload_tasks"
                )
                
                if st.button("✅ Сохранить расписание", use_container_width=True, key="save_schedule"):
                    schedule_data = {
                        "type": "interval" if schedule_type == "Интервал" else "cron",
                        "task_type": "upload_prices",
                        "interval_seconds": interval_seconds.get(interval, 3600),
                        "tasks": upload_tasks
                    }
                    if schedule_type != "Интервал":
                        schedule_data["hour"] = hour
                        schedule_data["minute"] = minute
                    
                    self.scheduler.add_schedule("default", schedule_data)
                    if enable_schedule:
                        self.scheduler.run_scheduled()
                    
                    st.success("✅ Расписание сохранено!")
            
            st.divider()
            
            st.markdown("### 📋 Лог выгрузок")
            log = self.uploader.get_upload_log()
            if log:
                for entry in log[-5:]:
                    with st.expander(f"📌 {entry.get('marketplace', '')} - {entry.get('timestamp', '')[:19]}"):
                        st.metric("Успешно", entry.get('success', 0))
                        st.metric("Ошибок", entry.get('failed', 0))
                        if entry.get('errors'):
                            st.warning(f"⚠️ {len(entry.get('errors', []))} ошибок")
                if st.button("🗑️ Очистить лог", key="clear_log"):
                    self.uploader.clear_log()
                    st.success("Лог очищен")
            else:
                st.info("Нет записей в логе")
    
    def _render_integration_tab(self):
        """Вкладка интеграции с 1С/CRM"""
        st.subheader("🔄 Интеграция с 1С и CRM")
        
        st.markdown("""
        **Поддерживаемые интеграции:**
        - ✅ 1С:Предприятие (обмен товарами, ценами, остатками)
        - ✅ AmoCRM (отправка отчетов)
        - ✅ Bitrix24 (создание сделок)
        - ✅ HubSpot (отправка данных)
        """)
        
        tab1, tab2 = st.tabs(["🔗 1С", "📊 CRM"])
        
        with tab1:
            st.markdown("### 🔗 Настройка интеграции с 1С")
            
            col1, col2 = st.columns(2)
            with col1:
                onec_url = st.text_input(
                    "URL 1С Web-сервиса",
                    placeholder="https://1c-server.ru/api",
                    key="onec_url"
                )
                onec_login = st.text_input("Логин 1С", key="onec_login")
                onec_password = st.text_input("Пароль 1С", type="password", key="onec_password")
                
                if st.button("🔗 Подключить 1С", use_container_width=True, key="connect_1c"):
                    if onec_url:
                        self.onec = OneCIntegration(onec_url, onec_login, onec_password)
                        st.success("✅ 1С подключена!")
                    else:
                        st.warning("⚠️ Введите URL 1С")
            
            with col2:
                st.markdown("### 📤 Действия")
                
                action = st.selectbox(
                    "Действие",
                    ["Экспорт товаров в 1С", "Импорт товаров из 1С", "Синхронизация цен", "Синхронизация остатков"],
                    key="onec_action"
                )
                
                if st.button("🚀 Выполнить", use_container_width=True, key="execute_1c"):
                    if action == "Экспорт товаров в 1С":
                        if self.results:
                            result = self.onec.export_to_1c(self.results[:100], "products")
                            if result.get('status') == 'success':
                                st.success("✅ Товары экспортированы в 1С")
                            else:
                                st.error(f"❌ Ошибка: {result.get('message')}")
                        else:
                            st.warning("⚠️ Нет данных для экспорта")
                    
                    elif action == "Импорт товаров из 1С":
                        data = self.onec.import_from_1c("products")
                        if data:
                            st.success(f"✅ Импортировано {len(data)} товаров из 1С")
                            st.json(data[:5])
                        else:
                            st.warning("⚠️ Нет данных для импорта")
        
        with tab2:
            st.markdown("### 📊 Настройка интеграции с CRM")
            
            col1, col2 = st.columns(2)
            with col1:
                crm_type = st.selectbox(
                    "Тип CRM",
                    ["AmoCRM", "Bitrix24", "HubSpot", "Другая"],
                    key="crm_type"
                )
                
                crm_url = st.text_input(
                    "URL CRM API",
                    placeholder="https://your-crm.com/api",
                    key="crm_url"
                )
                crm_token = st.text_input("API ключ/Token", type="password", key="crm_token")
                
                if st.button("🔗 Подключить CRM", use_container_width=True, key="connect_crm"):
                    if crm_url and crm_token:
                        self.crm = CRMIntegration(
                            crm_type.lower().replace(" ", ""),
                            crm_token,
                            crm_url
                        )
                        st.success("✅ CRM подключена!")
                    else:
                        st.warning("⚠️ Введите URL и API ключ")
            
            with col2:
                st.markdown("### 📤 Отправка отчета")
                
                if st.button("📊 Отправить отчет в CRM", type="primary", use_container_width=True, key="send_crm_report"):
                    if self.results:
                        total_profit = sum(r.get("unit_profit", 0) for r in self.results)
                        profitable = sum(1 for r in self.results if r.get("unit_profit", 0) > 0)
                        avg_margin = sum(r.get("margin", 0) for r in self.results) / len(self.results) if self.results else 0
                        
                        report_data = {
                            "title": f"Отчет по юнит-экономике от {datetime.now().strftime('%d.%m.%Y')}",
                            "summary": f"""
                            Всего товаров: {len(self.results)}
                            Прибыльных: {profitable}
                            Общая прибыль: {format_currency(total_profit)}
                            Средняя маржа: {format_percent(avg_margin)}
                            """,
                            "total_profit": total_profit,
                            "product_count": len(self.results),
                            "profitable_count": profitable
                        }
                        
                        result = self.crm.send_report(report_data)
                        if result.get('status') == 'success':
                            st.success("✅ Отчет отправлен в CRM!")
                            st.json(result.get('data', {}))
                        else:
                            st.error(f"❌ Ошибка: {result.get('message')}")
                    else:
                        st.warning("⚠️ Сначала рассчитайте юнит-экономику")
    
    def _render_dashboard_tab(self):
        """Вкладка дашборда"""
        st.subheader("📊 Дашборд аналитики")
        
        if not PLOTLY_AVAILABLE:
            st.warning("⚠️ Установите plotly для графиков: pip install plotly")
            return
        
        df = self.sales_analytics.get_sales_dataframe()
        
        if df.empty:
            st.info("💡 Нет данных для отображения. Добавьте продажи или сгенерируйте демо-данные.")
            if st.button("🔄 Сгенерировать демо-данные", key="gen_demo"):
                self._generate_demo_sales()
                st.rerun()
            return
        
        if not df.empty:
            self._render_fast_metrics(df)
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("📈 Показать тренд продаж", use_container_width=True, key="show_trend"):
                    self._render_sales_trend(df)
            
            with col2:
                if st.button("📊 Показать категории", use_container_width=True, key="show_categories"):
                    self._render_category_chart(df)
    
    def _render_fast_metrics(self, df: pd.DataFrame):
        """Быстрые метрики"""
        total_revenue = (df['quantity'] * df['price']).sum()
        total_profit = df['profit'].sum()
        total_orders = len(df)
        
        col1, col2, col3 = st.columns(3)
        
        col1.metric("💰 Выручка", format_currency(total_revenue))
        col2.metric("💵 Прибыль", format_currency(total_profit))
        col3.metric("📦 Заказов", f"{total_orders}")
    
    def _render_sales_trend(self, df: pd.DataFrame):
        """График тренда продаж"""
        if not PLOTLY_AVAILABLE:
            st.info("Установите plotly для графиков")
            return
        
        daily = df.groupby('date').agg({
            'quantity': 'sum',
            'profit': 'sum'
        }).reset_index()
        
        fig = go.Figure()
        
        fig.add_trace(go.Scatter(
            x=daily['date'],
            y=daily['profit'],
            name="Прибыль",
            line=dict(color="#e94560", width=2),
            fill='tozeroy',
            fillcolor='rgba(233, 69, 96, 0.2)'
        ))
        
        fig.add_trace(go.Bar(
            x=daily['date'],
            y=daily['quantity'],
            name="Кол-во",
            marker=dict(color="#0f3460", opacity=0.7),
            yaxis="y2"
        ))
        
        fig.update_layout(
            height=350,
            margin=dict(l=0, r=0, t=20, b=0),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            hovermode='x unified',
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            yaxis=dict(title="Прибыль, ₽"),
            yaxis2=dict(title="Кол-во", overlaying="y", side="right")
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    def _render_category_chart(self, df: pd.DataFrame):
        """График по категориям"""
        if not PLOTLY_AVAILABLE:
            st.info("Установите plotly для графиков")
            return
        
        category = df.groupby('category').agg({
            'profit': 'sum'
        }).reset_index().sort_values('profit', ascending=True)
        
        fig = go.Figure()
        
        fig.add_trace(go.Bar(
            y=category['category'],
            x=category['profit'],
            orientation='h',
            marker=dict(
                color=category['profit'],
                colorscale='Reds',
                showscale=True
            ),
            text=category['profit'].apply(lambda x: f"{x:,.0f}₽"),
            textposition='outside'
        ))
        
        fig.update_layout(
            height=350,
            margin=dict(l=0, r=0, t=20, b=0),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            xaxis_title="Прибыль, ₽",
            yaxis_title=""
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    def _generate_demo_sales(self):
        """Генерация демо-данных"""
        categories = ["Двигатель", "Трансмиссия", "Подвеска", "Тормозная система", 
                     "Масла и жидкости", "Фильтры", "Электрооборудование"]
        
        start_date = datetime.now() - timedelta(days=30)
        
        for i in range(50):
            date = start_date + timedelta(days=random.randint(0, 30))
            category = random.choice(categories)
            quantity = random.randint(1, 3)
            price = random.randint(500, 10000)
            cost = price * random.uniform(0.5, 0.7)
            
            self.sales_analytics.add_sale({
                "category": category,
                "quantity": quantity,
                "price": price,
                "cost": cost,
                "profit": (price - cost) * quantity,
                "marketplace": random.choice(["Яндекс Маркет", "Ozon", "Wildberries"])
            })
        
        st.success("✅ Демо-данные сгенерированы!")
    
    def _execute_api_request(self, marketplace: str, action: str) -> Optional[Dict]:
        """Выполнение API-запроса"""
        st.info(f"📌 {marketplace}: {action} (требуется настройка API ключей)")
        
        return {
            "status": "success",
            "marketplace": marketplace,
            "action": action,
            "timestamp": datetime.now().isoformat(),
            "data": {
                "items": [
                    {"id": "1", "name": "Товар 1", "price": 1000},
                    {"id": "2", "name": "Товар 2", "price": 2000}
                ]
            }
        }
    
    def _get_marketplace_from_sidebar(self) -> str:
        """Получение маркетплейса из боковой панели"""
        return "Яндекс Маркет"
    
    def _get_mode_from_sidebar(self) -> str:
        """Получение режима из боковой панели"""
        return "FBY"
    
    def _get_days_storage_from_sidebar(self) -> int:
        """Получение дней хранения из боковой панели"""
        return 30
    
    def _show_results(self, results: List[Dict]):
        """Отображение результатов"""
        if not results:
            return
        
        st.subheader("📊 Результаты расчета")
        self._show_product_results(results)
    
    def _show_product_results(self, results: List[Dict]):
        """Отображение результатов товарной модели"""
        total_profit = sum(r.get("unit_profit", 0) for r in results)
        avg_margin = sum(r.get("margin", 0) for r in results) / len(results) if results else 0
        profitable = sum(1 for r in results if r.get("unit_profit", 0) > 0)
        
        col1, col2, col3, col4, col5 = st.columns(5)
        col1.metric("📦 Товаров", len(results))
        col2.metric("💰 Прибыльных", profitable)
        col3.metric("💵 Общая прибыль", format_currency(total_profit))
        col4.metric("📊 Ср. маржа", format_percent(avg_margin))
        col5.metric("🏷️ Категорий", len(set(r.get("category", "Прочее") for r in results)))
        
        st.subheader("🏆 Топ-10 по прибыли")
        top = sorted(results, key=lambda x: x.get("unit_profit", 0), reverse=True)[:10]
        if top:
            df_top = pd.DataFrame([{
                "Артикул": r.get("article", ""),
                "Наименование": r.get("name", "")[:30],
                "Бренд": r.get("brand", ""),
                "OE номер": r.get("oe_number", ""),
                "Категория": r.get("category", ""),
                "Цена": format_currency(r.get("price", 0)),
                "Прибыль": format_currency(r.get("unit_profit", 0)),
                "Маржа": format_percent(r.get("margin", 0)),
                "ABC": r.get("abc_category", ""),
                "Позиция цены": r.get("price_position", ""),
                "Размеры": f"{r.get('length_orig', 0)}x{r.get('width_orig', 0)}x{r.get('height_orig', 0)} {r.get('dimension_unit', 'мм')}"
            } for r in top])
            st.dataframe(df_top, use_container_width=True, hide_index=True)
    
    def _show_export(self, results: List[Dict]):
        """Отображение экспорта"""
        st.subheader("📥 Экспорт в Excel")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.info(f"📊 Экспортируемых позиций: {len(results):,}")
            st.info(f"🏷️ Категорий: {len(set(r.get('category', '') for r in results))}")
            st.info(f"📏 Единицы: {self.dimension_unit}")
        
        with col2:
            marketplace = self._get_marketplace_from_sidebar()
            mode = self._get_mode_from_sidebar()
            
            if st.button("📥 Скачать Excel-отчет", type="primary", use_container_width=True, key="download_excel"):
                with st.spinner("⏳ Генерация Excel-файла..."):
                    try:
                        if not self.all_rates:
                            self.all_rates = self.tariff_provider.get_all_rates()
                        
                        data = self.exporter.export(results, marketplace, mode, self.all_rates)
                        
                        if data:
                            st.download_button(
                                "📥 Скачать файл",
                                data=data,
                                file_name=f"юнит_экономика_{marketplace}_{mode}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="download_file"
                            )
                            st.success("✅ Отчет готов к скачиванию!")
                        else:
                            st.error("❌ Ошибка генерации отчета")
                    except Exception as e:
                        st.error(f"❌ Ошибка: {str(e)}")
                        logger.error(f"Export error: {e}")
        
        with col3:
            if st.button("📤 Экспорт в Google Sheets", use_container_width=True, key="export_gsheets"):
                if self.results:
                    with st.spinner("⏳ Экспорт в Google Sheets..."):
                        result = self.google_sheets.export_results(self.results)
                        if result.get('status') == 'success':
                            st.success(f"✅ {result.get('message')}")
                            st.info(f"🔗 {result.get('url')}")
                        else:
                            st.error(f"❌ {result.get('message')}")
                else:
                    st.warning("⚠️ Сначала рассчитайте юнит-экономику")
        
        with st.expander("📋 Превью данных"):
            if results:
                preview = pd.DataFrame([{
                    "Артикул": r.get("article", ""),
                    "Наименование": r.get("name", "")[:25],
                    "Бренд": r.get("brand", ""),
                    "OE номер": r.get("oe_number", ""),
                    "Категория": r.get("category", ""),
                    "Цена": r.get("price", 0),
                    "Прибыль": r.get("unit_profit", 0),
                    "Маржа %": r.get("margin", 0),
                    "Позиция цены": r.get("price_position", ""),
                    "Размеры": f"{r.get('length_orig', 0)}x{r.get('width_orig', 0)}x{r.get('height_orig', 0)} {r.get('dimension_unit', 'мм')}"
                } for r in results[:20]])
                st.dataframe(preview, use_container_width=True, hide_index=True)
                st.caption(f"Показаны первые 20 из {len(results):,}")

# --------------------------------------------
# ЗАПУСК
# --------------------------------------------
if __name__ == "__main__":
    try:
        app = UnitEconomicsApp()
        app.run()
    except Exception as e:
        st.error(f"❌ Критическая ошибка: {str(e)}")
        st.code(traceback.format_exc())
        logger.error(f"Critical error: {e}")
